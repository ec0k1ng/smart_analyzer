from __future__ import annotations

import ast
import builtins
import importlib.util
import json
import os
import re
from copy import deepcopy
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection

CONFIG_DIR = Path(__file__).resolve().parent
RULE_SPECS_DIR = CONFIG_DIR / "rule_specs"
KPI_SPECS_DIR = CONFIG_DIR / "kpi_specs"
DERIVED_SIGNALS_DIR = CONFIG_DIR / "derived_signals"
REPORT_TEMPLATES_DIR = CONFIG_DIR / "report_templates"
FORMULA_SIGNALS_PATH = CONFIG_DIR / "signal_library_formulas.json"
KPI_GROUPS_PATH = CONFIG_DIR / "kpi_groups.json"
CHART_VIEW_STATE_PATH = CONFIG_DIR / "chart_view_state.json"
LEGACY_RULE_DEFINITIONS_PATH = CONFIG_DIR / "rule_definitions.json"
LEGACY_KPI_DEFINITIONS_PATH = CONFIG_DIR / "kpi_definitions.json"
LEGACY_INTERFACE_MAPPING_PATH = CONFIG_DIR / "interface_mapping.json"
INTERFACE_MAPPING_PATH = CONFIG_DIR / "interface_mapping.xlsx"
SYSTEM_MAPPING_SHEET = "系统信号"
CUSTOM_MAPPING_SHEET = "自定义信号"
GUIDE_SHEET = "说明"
REFERENCE_SHEET = "参考信息"
INTERFACE_MAPPING_METADATA_SHEET = "_metadata"
DEFAULT_INTERFACE_ACTUAL_NAME_COLUMN_COUNT = 3


def build_system_interface_mapping_headers(actual_name_column_count: int = DEFAULT_INTERFACE_ACTUAL_NAME_COLUMN_COUNT) -> list[str]:
    count = max(DEFAULT_INTERFACE_ACTUAL_NAME_COLUMN_COUNT, int(actual_name_column_count or DEFAULT_INTERFACE_ACTUAL_NAME_COLUMN_COUNT))
    return ["raw_input_name", "description", "from", *[f"actual_signal_name_{index}" for index in range(1, count + 1)]]


def build_custom_interface_mapping_headers(actual_name_column_count: int = DEFAULT_INTERFACE_ACTUAL_NAME_COLUMN_COUNT) -> list[str]:
    count = max(DEFAULT_INTERFACE_ACTUAL_NAME_COLUMN_COUNT, int(actual_name_column_count or DEFAULT_INTERFACE_ACTUAL_NAME_COLUMN_COUNT))
    return ["raw_input_name", *[f"actual_signal_name_{index}" for index in range(1, count + 1)]]

RAW_INPUT_DESCRIPTIONS = {
    "time_s": "时间轴，单位 s。",
    "vehicle_speed_kph": "车速，单位 kph。",
    "wheel_speed_fl_kph": "左前轮轮速，单位 kph。",
    "wheel_speed_fr_kph": "右前轮轮速，单位 kph。",
    "wheel_speed_rl_kph": "左后轮轮速，单位 kph。",
    "wheel_speed_rr_kph": "右后轮轮速，单位 kph。",
    "longitudinal_accel_mps2": "纵向加速度，单位 m/s^2。",
    "yaw_rate_degps": "横摆角速度，单位 deg/s。",
    "steering_wheel_angle_deg": "方向盘转角，单位 deg。",
    "accel_pedal_pct": "油门开度，单位 %。",
    "brake_depth_pct": "制动深度，单位 %。",
    "torque_request_nm": "请求扭矩，单位 Nm。",
    "torque_actual_nm": "实际扭矩，单位 Nm。",
    "brake_pressure_fl_bar": "左前制动压力，单位 bar。",
    "brake_pressure_fr_bar": "右前制动压力，单位 bar。",
    "tcs_active": "TCS 总激活标志，布尔/0-1。",
    "tcs_active_fl": "左前轮 TCS 激活标志，布尔/0-1。",
    "tcs_active_fr": "右前轮 TCS 激活标志，布尔/0-1。",
    "tcs_active_rl": "左后轮 TCS 激活标志，布尔/0-1。",
    "tcs_active_rr": "右后轮 TCS 激活标志，布尔/0-1。",
    "abs_active": "ABS 总激活标志，布尔/0-1。",
    "abs_active_fl": "左前轮 ABS 激活标志，布尔/0-1。",
    "abs_active_fr": "右前轮 ABS 激活标志，布尔/0-1。",
    "abs_active_rl": "左后轮 ABS 激活标志，布尔/0-1。",
    "abs_active_rr": "右后轮 ABS 激活标志，布尔/0-1。",
}

DEFAULT_RULE_TEMPLATE = '''from __future__ import annotations

DISPLAY_NAME = "{display_name}"
IS_TEMPLATE = False

RULE_DEFINITION = {{
    "rule_id": "{rule_id}",
    "category": "performance",
    "title": "{display_name}",
    "trend_source": "{kpi_name}",
    "enabled": True,
    "raw_inputs": ["time_s", "vehicle_speed"],
    "feature_inputs": [],
    "threshold": 0.0,
    "unit": "",
    "source": "manual",
}}


def evaluate_rule(dataframe, scenarios, scenario, settings):
    measured_value = 0.0
    threshold_value = settings.get_rule_threshold(RULE_DEFINITION["rule_id"], RULE_DEFINITION["threshold"])
    passed = measured_value <= threshold_value
    return {{
        "measured_value": measured_value,
        "threshold_value": threshold_value,
        "status": "pass" if passed else "fail",
        "severity": "info" if passed else "high",
        "message": "请根据业务规则补充算法逻辑",
        "threshold_source": settings.get_rule_threshold_source(RULE_DEFINITION["rule_id"], RULE_DEFINITION["source"]),
        "confidence": 1.0,
    }}
'''

DEFAULT_KPI_TEMPLATE = '''from __future__ import annotations

IS_TEMPLATE = False

KPI_DEFINITION = {{
    "name": "{kpi_name}",
    "title": "{title}",
    "raw_inputs": [
        "time_s",  # 时间轴，单位 s
        "vehicle_speed_kph",  # 车速，单位 kph
    ],
    "derived_inputs": [],
    "trend_source": "{kpi_name}",
    "unit": "",
    "description": "说明这个 KPI 计算的业务含义",
    "algorithm_summary": "用文字概述算法思路、关键公式和边界处理方式",
    "threshold": 0.0,
    "source": "kpi_definition",
    "pass_condition": "value <= threshold",
    "rule_description": "说明该 KPI 需要满足什么条件才算达标",
    "pass_message": "该 KPI 达标",
    "fail_message": "该 KPI 未达标",
}}

CALIBRATION = {{
    "missing_value_fill": 0.0,  # 算法标定量统一放在这里；命名建议使用 snake_case，并尽量带上业务含义或单位后缀
}}


def calculate_kpi(dataframe):
    return 0.0


def calculate_kpi_series(dataframe):
    return dataframe["vehicle_speed_kph"].fillna(CALIBRATION["missing_value_fill"])
'''

DEFAULT_DERIVED_SIGNAL_TEMPLATE = '''from __future__ import annotations

IS_TEMPLATE = False

import pandas as pd


DERIVED_SIGNAL_DEFINITION = {{
    "name": "{signal_name}",
    "title": "{title}",
    "raw_inputs": [
        "time_s",  # 时间轴，单位 s
        "vehicle_speed_kph",  # 车速，单位 kph
    ],
    "derived_inputs": [],
    "description": "说明这个派生量的业务含义，以及哪些 KPI 会复用它",
    "algorithm_summary": "用文字概述算法思路、关键公式和边界处理方式",
}}

CALIBRATION = {{
    "missing_fill_value": 0.0,  # 算法标定量统一放在这里；命名建议使用 snake_case，并尽量带上业务含义或单位后缀
}}


def calculate_signal(dataframe):
    return pd.to_numeric(dataframe["vehicle_speed_kph"], errors="coerce").fillna(CALIBRATION["missing_fill_value"])
'''

DEFAULT_DERIVED_SIGNAL_GUIDE_TEMPLATE = """GUIDE_TEXT = '''派生量示例与详细讲解

用途：
1. 这是 UI 下拉菜单中的第一个选项，用于给用户查看派生量格式。
2. 这不是实际执行派生量，分析引擎会自动跳过该文件。
3. 派生量用于承载多个 KPI 共用、且希望只计算一次的中间序列。

给 AI 的格式化要求：
请为 TCS Smart Analyzer 生成一个派生量 Python 文件，严格按下面格式输出，不要解释：

from __future__ import annotations

IS_TEMPLATE = False

import pandas as pd

DERIVED_SIGNAL_DEFINITION = {
    "name": "唯一派生量名称，例如 slip_kph。必须使用英文，因为它会作为信号名参与依赖声明、绘图和计算引用",
    "title": "派生量标题。请使用中文，供界面标签、下拉框和结果说明展示",
    "raw_inputs": [
        "time_s",  # 示例：时间轴，单位 s
        "vehicle_speed_kph",  # 示例：车速，单位 kph
    ],
    "derived_inputs": ["如果依赖其他派生量，在这里列出，可为空"],
    "description": "说明这个派生量代表什么，以及哪些 KPI 会复用它",
    "algorithm_summary": "用文字概述算法思路、关键公式和边界处理方式",
}

CALIBRATION = {
    "calibration_name": 0.0,  # 所有可调标定量统一集中在这里；请使用 snake_case，并在右侧写清楚注释
}

def calculate_signal(dataframe):
    return pd.Series(...)

说明：
- 对用户来说，raw_inputs 就是唯一需要维护的输入声明；接口映射表第一列会自动由所有 KPI 和派生量的 raw_inputs 汇总同步，不需要再额外维护另一份“必需信号清单”。
- raw_inputs 里的标准输入名称应带单位，例如 vehicle_speed_kph、yaw_rate_degps；布尔状态量保持语义名即可。
- raw_inputs 建议逐行书写并在右侧补注释，明确物理意义与单位，避免接口映射和算法理解歧义。
- 派生量通常应输出一个与数据长度一致、随时间连续变化的过程曲线；只有当该量明确表示“整个文件的固有标量属性”时，才允许返回全时段同一数值的水平直线。
- 如果你认为派生量应该输出水平直线，必须先和用户确认它到底是“单值标量属性”还是“实时连续变化量”；未确认前不允许自作主张。
- calculate_signal(dataframe) 必须返回一个与数据长度一致的序列，曲线界面会直接按 DERIVED_SIGNAL_DEFINITION["name"] 显示这条派生量曲线。
- 和 KPI 不同，派生量不需要 trend_source，也不需要额外再写一个 calculate_kpi_series；calculate_signal(dataframe) 的返回值本身就是曲线来源。
- 派生量没有单独的 trend_source 字段；它写回分析数据表后的列名，就是 DERIVED_SIGNAL_DEFINITION["name"] 本身，所以 name 必须稳定、唯一、可直接用于绘图。
- 派生量只负责输出一个与数据长度一致的序列，供多个 KPI 复用。
- 派生量会在单个文件分析过程中只计算一次，然后写回分析数据表。
- 如果多个 KPI 共用一个中间量，应优先建派生量，不要在多个 KPI 文件里重复算。
- raw_inputs 会自动进入接口映射 Excel 第一列来源统计，所以一定要写全。
- derived_inputs 用于声明派生量之间的依赖，分析引擎会按依赖顺序自动计算。
- 所有可调算法参数都应集中放在 CALIBRATION 区块，放在定义区后、函数前；不要把标定量零散写在文件各处。
- CALIBRATION 里的键名请使用 snake_case，并尽量把物理意义、工况或单位写进名字；每个条目右侧都要写中文注释。
- 如果派生量依赖其他派生量，请把被依赖项完整写进 derived_inputs，格式与 KPI 文件中的 derived_inputs 保持一致。
- 当前清单中的每一行都已压缩展示；连字符后面的说明优先取算法概述，没有算法概述时才回退到 description。
- 如果某个派生量只在局部工况下有业务意义，建议在 description 和 algorithm_summary 里写清楚非有效区间是返回 0.0、保持上一值、NaN 还是其它占位值；这不是硬编码格式要求，但最好提前讲明，避免曲线含义不清。

当前可用派生量清单：
{{DERIVED_CATALOG}}

当前接口映射表中的标准输入量：
{{RAW_INPUT_CATALOG}}
'''

IS_TEMPLATE = True
"""

DEFAULT_REPORT_TEMPLATE = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <title>{{ report_title }}</title>
  <style>
    body { font-family: "Microsoft YaHei", sans-serif; margin: 24px; color: #1f2937; }
    h1, h2 { color: #155e75; }
    .hero { padding: 16px 18px; background: linear-gradient(135deg, #ecfeff, #e0f2fe); border: 1px solid #bae6fd; margin-bottom: 18px; }
    .cards { display: flex; gap: 12px; flex-wrap: wrap; margin-bottom: 18px; }
    .card { min-width: 180px; padding: 12px; border: 1px solid #dbeafe; background: #f8fafc; }
    table { border-collapse: collapse; width: 100%; margin-bottom: 20px; }
    th, td { border: 1px solid #cbd5e1; padding: 8px 10px; font-size: 13px; }
    th { background: #f0f9ff; text-align: left; }
  </style>
</head>
<body>
  <div class="hero">
    <h1>{{ report_title }}</h1>
        <p>汇总文件数：{{ file_count }}</p>
        <p>分析配置：{{ analysis_profile }}</p>
    <p>分析摘要：{{ report_summary }}</p>
    <p>生成时间：{{ generated_at }}</p>
  </div>

  <div class="cards">
        <div class="card"><strong>文件数</strong><div>{{ file_count }}</div></div>
        <div class="card"><strong>KPI 结果数</strong><div>{{ kpis|length }}</div></div>
        <div class="card"><strong>规则数</strong><div>{{ rules|length }}</div></div>
        <div class="card"><strong>分组条目数</strong><div>{{ results|length }}</div></div>
  </div>

    {% for item in results %}
    <section style="margin-bottom: 22px; border: 1px solid #dbeafe; border-radius: 14px; overflow: hidden; background: #ffffff;">
        <div style="padding: 10px 14px; background: #e0f2fe; color: #0c4a6e; font-weight: 700;">
            {{ item.group_name }}
        </div>
        <div style="padding: 10px 14px; background: #f8fafc; color: #334155; font-weight: 600; border-top: 1px solid #e2e8f0;">
            {{ item.source_name }}
        </div>
        <table style="margin-bottom: 0;">
            <thead><tr><th>名称</th><th>说明</th><th>单位</th><th>数值</th><th>规则</th><th>结果</th></tr></thead>
            <tbody>
            {% for kpi in item.kpis %}
                <tr><td>{{ kpi.title }}</td><td>{{ kpi.description }}</td><td>{{ kpi.unit }}</td><td>{{ kpi.value_text }}</td><td>{{ kpi.rule_description }}</td><td>{{ kpi.result_label }}</td></tr>
            {% endfor %}
            </tbody>
        </table>
    </section>
    {% endfor %}
</body>
</html>
'''

DEFAULT_REPORT_GUIDE_TEMPLATE = '''<!--
报告模板示例与详细讲解

用途：
1. 这是报告模板工作台中的第一个选项，用于给用户查看模板格式。
2. 这个文件本身也可以作为最小可运行模板。
3. 你可以把本文件和下面的 AI 格式化要求一起发给 AI，让 AI 直接生成你的报告模板。
4. 如果你本地已经有 Word 报告模板，也可以把 Word 模板的版式要求完整描述给 AI，让 AI 翻译成 HTML 模板代码。

给 AI 的格式化要求：
请为 TCS Smart Analyzer 生成一个 HTML 报告模板，严格输出一个可直接保存为 .html 的 Jinja2 模板文件，不要解释。

模板中可直接使用的变量：
- report_title: 报告标题字符串。
- source_path: 当前报告来源路径；多文件汇总报告时固定为“多文件汇总报告”。
- source_name: 当前报告来源文件名；多文件汇总报告时通常为空字符串。
- source_stem: 当前报告来源文件名去掉扩展名后的部分；多文件汇总报告时通常为空字符串。
- analysis_profile: 当前分析配置名称。
- generated_at: 本次报告生成时间。
- report_summary: 分析摘要文本。
- mapped_columns: 标准信号到实际信号名的映射字典。
- metadata: 报告级元数据字典，多文件汇总时包含 analysis_profiles、file_count、config_paths、group_keys 等信息。
- file_count: 当前报告汇总的文件数量。
- results: 文件级结果列表；每项都包含 source_path、source_name、source_stem、analysis_profile、generated_at、mapped_columns、metadata、group_key、group_name、kpi_count、kpis、rules。
- files: 与 results 相同，提供给喜欢用 files 这个名字的模板作者。
- kpis: 扁平化 KPI 结果列表；每项额外包含 source_path、source_name、source_stem、group_key、group_name，可用于直接生成跨文件总表。
- rules: 扁平化规则结果列表；每项额外包含 source_path、source_name、group_key。

常用循环示例：
- {% for item in results %} ... {{ item.source_name }} ... {% endfor %}
- {% for item in kpis %} ... {{ item.source_name }} / {{ item.title }} ... {% endfor %}

建议要求：
- 页面必须是完整 HTML
- 保留中文展示
- 保证没有变量时也能正常渲染
- KPI 结果至少展示一张表
-->
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="utf-8">
    <title>{{ report_title }}</title>
    <style>
        body { font-family: "Microsoft YaHei", sans-serif; margin: 24px; color: #1f2937; }
        h1, h2 { color: #155e75; }
        .hero { padding: 16px 18px; background: linear-gradient(135deg, #ecfeff, #e0f2fe); border: 1px solid #bae6fd; margin-bottom: 18px; }
        table { border-collapse: collapse; width: 100%; margin-bottom: 20px; }
        th, td { border: 1px solid #cbd5e1; padding: 8px 10px; font-size: 13px; }
        th { background: #f0f9ff; text-align: left; }
    </style>
</head>
<body>
    <div class="hero">
        <h1>{{ report_title }}</h1>
        <p>汇总文件数：{{ file_count }}</p>
        <p>分析配置：{{ analysis_profile }}</p>
        <p>分析摘要：{{ report_summary }}</p>
        <p>生成时间：{{ generated_at }}</p>
    </div>

    {% for item in results %}
    <section style="margin-bottom: 22px; border: 1px solid #dbeafe; border-radius: 14px; overflow: hidden; background: #ffffff;">
        <div style="padding: 10px 14px; background: #e0f2fe; color: #0c4a6e; font-weight: 700;">
            {{ item.group_name }}
        </div>
        <div style="padding: 10px 14px; background: #f8fafc; color: #334155; font-weight: 600; border-top: 1px solid #e2e8f0;">
            {{ item.source_name }}
        </div>
        <table style="margin-bottom: 0;">
            <thead><tr><th>名称</th><th>说明</th><th>单位</th><th>数值</th><th>规则</th><th>结果</th></tr></thead>
            <tbody>
            {% for kpi in item.kpis %}
                <tr><td>{{ kpi.title }}</td><td>{{ kpi.description }}</td><td>{{ kpi.unit }}</td><td>{{ kpi.value_text }}</td><td>{{ kpi.rule_description }}</td><td>{{ kpi.result_label }}</td></tr>
            {% endfor %}
            </tbody>
        </table>
    </section>
    {% endfor %}
</body>
</html>
'''

KPI_GUIDE_TEMPLATE = '''GUIDE_TEXT = """KPI 示例与详细讲解

用途：
1. 这是 UI 下拉菜单中的第一个选项，用于给用户查看新版 KPI 格式。
2. 这不是实际执行 KPI，分析引擎会自动跳过该文件。
3. 现在不再单独维护规则文件。每一项 KPI 自己同时定义“数值如何算”和“结果如何判定”。

给 AI 的格式化要求：
请为 TCS Smart Analyzer 生成一个 KPI Python 文件，严格按下面格式输出，不要解释：

from __future__ import annotations

IS_TEMPLATE = False

KPI_DEFINITION = {
    "name": "唯一KPI名称，例如 max_slip_kph。必须使用英文，因为它会作为KPI信号名参与依赖、绘图和结果索引",
    "title": "KPI标题。请使用中文，供界面标签、结果表和报告展示",
    "raw_inputs": [
        "time_s",  # 示例：时间轴，单位 s
        "vehicle_speed_kph",  # 示例：车速，单位 kph
    ],
    "derived_inputs": ["如果依赖派生量，在这里列出，例如 slip_kph"],
    "trend_source": "必须与 name 完全一致，例如 max_slip_kph。曲线界面的 KPI 信号只认 KPI 自己的 name",
    "unit": "单位",
    "description": "说明这个 KPI 算出来代表什么",
    "algorithm_summary": "用文字概述算法思路、关键公式和边界处理方式",
    "threshold": 0.0,
    "source": "阈值来源说明",
    "pass_condition": "value <= threshold",
    "rule_description": "说明该 KPI 需要满足什么条件才算达标",
    "pass_message": "该 KPI 达标时显示的话",
    "fail_message": "该 KPI 未达标时显示的话",
}

CALIBRATION = {
    "calibration_name": 0.0,  # 所有可调标定量统一集中在这里；请使用 snake_case，并在右侧写清楚注释
}

def calculate_kpi(dataframe):
    return 0.0

def calculate_kpi_series(dataframe):
    return dataframe["某个连续过程序列"]

说明：
- 对用户来说，raw_inputs 是唯一需要维护的输入声明；接口映射表第一列会自动由所有 KPI 和派生量的 raw_inputs 汇总同步，不需要再额外维护另一份“必需信号清单”。
- raw_inputs 里的标准输入名称应带单位，例如 vehicle_speed_kph、wheel_speed_fl_kph、yaw_rate_degps。
- raw_inputs 建议逐行书写并在右侧补注释，明确物理意义与单位。
- calculate_kpi(dataframe)：功能实现字段，负责用 Python 算出 KPI 数值。
- calculate_kpi_series(dataframe)：必填字段，必须返回与数据长度一致的连续过程曲线，供曲线界面实时检查 KPI 算法过程和峰值位置。
- algorithm_summary：必填字段，用一句到几句中文讲清楚算法核心逻辑、关键公式和边界处理。
- 不要输出 DISPLAY_NAME，也不要把用户新建 KPI 文件写成 IS_TEMPLATE = True；这两个字段只属于系统示例文件，不属于实际 KPI 定义。
- 如果多个 KPI 共用同一个中间量，应优先把它抽成派生量，并在 derived_inputs 中显式声明依赖。
- 如果某个中间量只被当前 KPI 使用，也可以继续直接写在 KPI 文件内，保持算法透明可改。
- trend_source 不是可选项，必须填写为 KPI_DEFINITION["name"] 本身；曲线界面展示 KPI 过程时，读取的是 KPI 名称对应的连续序列。
- 如果想在曲线里看到 RMS 包络、累计最大值、滑窗结果等过程，就应在 calculate_kpi_series 中直接返回该过程序列；不要把 trend_source 写成其它临时列名。
- 所有可调算法参数都应集中放在 CALIBRATION 区块，放在 KPI_DEFINITION 之后、函数之前；不要把标定量零散写在文件顶部和函数内部。
- CALIBRATION 里的键名请使用 snake_case，并尽量把物理意义、工况或单位写进名字；每个条目右侧都要写中文注释。
- pass_condition：通过判断字段，使用 Python 表达式；可直接引用 value、threshold、dataframe、np、pd、math、source_path、source_name、source_stem、analysis_profile、generated_at、mapped_columns。
- rule_description：展示给工程人员看的规则描述。
- dataframe 是标准化后的分析数据表，里面既包含原始标准信号，也包含当前分析所需、已经按依赖顺序算好的派生量。
- dataframe.attrs["source_name"]：当前分析文件名，例如 demo.csv。
- dataframe.attrs["source_stem"]：当前分析文件名去掉扩展名后的值，例如 demo。
- dataframe.attrs["source_path"]：当前分析文件绝对路径字符串。
- dataframe.attrs["analysis_profile"]：当前分析配置名称。
- dataframe.attrs["generated_at"]：本次分析生成时间。
- dataframe.attrs["mapped_columns"]：标准信号到实际信号名的映射字典。
- 若你要让 AI 同时设计 KPI 与报告模板，还要明确告诉它：报告模板阶段可额外使用 report_title、metadata、kpis、rules、results、files、file_count 等变量。
- raw_inputs 会自动进入接口映射 Excel 第一列来源统计，所以一定要写全。
- 当前清单中的每一行都已压缩展示；连字符后面的说明优先取算法概述，没有算法概述时才回退到 description。
- derived_inputs 不会直接参与接口映射，但它引用的派生量 raw_inputs 也会自动进入接口映射来源统计。

当前可用派生量清单：
{{DERIVED_CATALOG}}

当前接口映射表中的标准输入量：
{{RAW_INPUT_CATALOG}}
"""

IS_TEMPLATE = True
'''

_PYTHON_BUILTIN_NAMES = set(dir(builtins))


@dataclass(slots=True)
class ConfigValidationIssue:
    path: Path
    message: str
    line: int | None = None
    column: int | None = None
    code: str = ""


def _extract_target_names(target) -> set[str]:  # noqa: ANN001
    if isinstance(target, ast.Name):
        return {target.id}
    if isinstance(target, (ast.Tuple, ast.List)):
        names: set[str] = set()
        for item in target.elts:
            names.update(_extract_target_names(item))
        return names
    if isinstance(target, ast.Starred):
        return _extract_target_names(target.value)
    return set()


class _ScopeSeedCollector(ast.NodeVisitor):
    def __init__(self) -> None:
        self.bound_names: set[str] = set()

    def _visit_statement_list(self, statements: list[ast.stmt]) -> None:
        for statement in statements:
            self.visit(statement)

    def visit_FunctionDef(self, node: ast.FunctionDef) -> None:  # noqa: N802
        self.bound_names.add(node.name)

    def visit_AsyncFunctionDef(self, node: ast.AsyncFunctionDef) -> None:  # noqa: N802
        self.bound_names.add(node.name)

    def visit_ClassDef(self, node: ast.ClassDef) -> None:  # noqa: N802
        self.bound_names.add(node.name)

    def visit_Import(self, node: ast.Import) -> None:  # noqa: N802
        for alias in node.names:
            self.bound_names.add(alias.asname or alias.name.split(".", 1)[0])

    def visit_ImportFrom(self, node: ast.ImportFrom) -> None:  # noqa: N802
        for alias in node.names:
            if alias.name == "*":
                continue
            self.bound_names.add(alias.asname or alias.name)

    def visit_Assign(self, node: ast.Assign) -> None:  # noqa: N802
        for target in node.targets:
            self.bound_names.update(_extract_target_names(target))
        self.generic_visit(node.value)

    def visit_AnnAssign(self, node: ast.AnnAssign) -> None:  # noqa: N802
        self.bound_names.update(_extract_target_names(node.target))
        if node.value is not None:
            self.generic_visit(node.value)

    def visit_AugAssign(self, node: ast.AugAssign) -> None:  # noqa: N802
        self.bound_names.update(_extract_target_names(node.target))
        self.generic_visit(node.value)

    def visit_For(self, node: ast.For) -> None:  # noqa: N802
        self.bound_names.update(_extract_target_names(node.target))
        self.visit(node.iter)
        self._visit_statement_list(node.body)
        self._visit_statement_list(node.orelse)

    def visit_AsyncFor(self, node: ast.AsyncFor) -> None:  # noqa: N802
        self.bound_names.update(_extract_target_names(node.target))
        self.visit(node.iter)
        self._visit_statement_list(node.body)
        self._visit_statement_list(node.orelse)

    def visit_If(self, node: ast.If) -> None:  # noqa: N802
        self.visit(node.test)
        self._visit_statement_list(node.body)
        self._visit_statement_list(node.orelse)

    def visit_While(self, node: ast.While) -> None:  # noqa: N802
        self.visit(node.test)
        self._visit_statement_list(node.body)
        self._visit_statement_list(node.orelse)

    def visit_With(self, node: ast.With) -> None:  # noqa: N802
        for item in node.items:
            if item.optional_vars is not None:
                self.bound_names.update(_extract_target_names(item.optional_vars))
            self.visit(item.context_expr)
        self._visit_statement_list(node.body)

    def visit_AsyncWith(self, node: ast.AsyncWith) -> None:  # noqa: N802
        for item in node.items:
            if item.optional_vars is not None:
                self.bound_names.update(_extract_target_names(item.optional_vars))
            self.visit(item.context_expr)
        self._visit_statement_list(node.body)

    def visit_ExceptHandler(self, node: ast.ExceptHandler) -> None:  # noqa: N802
        if node.name:
            self.bound_names.add(str(node.name))
        if node.type is not None:
            self.visit(node.type)
        self._visit_statement_list(node.body)

    def visit_Try(self, node: ast.Try) -> None:  # noqa: N802
        self._visit_statement_list(node.body)
        for handler in node.handlers:
            self.visit(handler)
        self._visit_statement_list(node.orelse)
        self._visit_statement_list(node.finalbody)


def _collect_scope_seed_names(body: list[ast.stmt]) -> set[str]:
    collector = _ScopeSeedCollector()
    for statement in body:
        collector.visit(statement)
    return collector.bound_names


def _function_argument_names(arguments: ast.arguments) -> set[str]:
    names = [*arguments.posonlyargs, *arguments.args, *arguments.kwonlyargs]
    if arguments.vararg is not None:
        names.append(arguments.vararg)
    if arguments.kwarg is not None:
        names.append(arguments.kwarg)
    return {item.arg for item in names}


class _UndefinedNameValidator(ast.NodeVisitor):
    def __init__(self, path: Path) -> None:
        self.path = path
        self.issues: list[ConfigValidationIssue] = []
        self._scopes: list[set[str]] = []
        self._reported_locations: set[tuple[int, int, str]] = set()

    def _push_scope(self, names: set[str]) -> None:
        self._scopes.append(set(names))

    def _pop_scope(self) -> None:
        self._scopes.pop()

    def _is_defined(self, name: str) -> bool:
        return any(name in scope for scope in reversed(self._scopes))

    def _report_undefined(self, node: ast.Name) -> None:
        line = getattr(node, "lineno", None)
        column = getattr(node, "col_offset", None)
        if line is None or column is None:
            return
        key = (line, column, node.id)
        if key in self._reported_locations:
            return
        self._reported_locations.add(key)
        self.issues.append(
            ConfigValidationIssue(
                path=self.path,
                message=f"未定义名称: {node.id}",
                line=int(line),
                column=int(column) + 1,
                code="undefined-name",
            )
        )

    def visit_Module(self, node: ast.Module) -> None:  # noqa: N802
        self._push_scope(_collect_scope_seed_names(node.body) | _PYTHON_BUILTIN_NAMES)
        for statement in node.body:
            self.visit(statement)
        self._pop_scope()

    def visit_FunctionDef(self, node: ast.FunctionDef) -> None:  # noqa: N802
        for decorator in node.decorator_list:
            self.visit(decorator)
        for default in [*node.args.defaults, *(default for default in node.args.kw_defaults if default is not None)]:
            self.visit(default)
        if node.returns is not None:
            self.visit(node.returns)
        self._push_scope(_collect_scope_seed_names(node.body) | _function_argument_names(node.args) | _PYTHON_BUILTIN_NAMES)
        for statement in node.body:
            self.visit(statement)
        self._pop_scope()

    def visit_AsyncFunctionDef(self, node: ast.AsyncFunctionDef) -> None:  # noqa: N802
        self.visit_FunctionDef(node)

    def visit_Lambda(self, node: ast.Lambda) -> None:  # noqa: N802
        for default in [*node.args.defaults, *(default for default in node.args.kw_defaults if default is not None)]:
            self.visit(default)
        self._push_scope(_function_argument_names(node.args) | _PYTHON_BUILTIN_NAMES)
        self.visit(node.body)
        self._pop_scope()

    def visit_ClassDef(self, node: ast.ClassDef) -> None:  # noqa: N802
        for decorator in node.decorator_list:
            self.visit(decorator)
        for base in node.bases:
            self.visit(base)
        for keyword in node.keywords:
            self.visit(keyword)
        self._push_scope(_collect_scope_seed_names(node.body) | {"__module__", "__qualname__"} | _PYTHON_BUILTIN_NAMES)
        for statement in node.body:
            self.visit(statement)
        self._pop_scope()

    def _visit_comprehension(self, generators: list[ast.comprehension], visit_value) -> None:
        bound_names: set[str] = set()
        self._push_scope(bound_names | _PYTHON_BUILTIN_NAMES)
        for generator in generators:
            self.visit(generator.iter)
            generator_names = _extract_target_names(generator.target)
            bound_names.update(generator_names)
            self._scopes[-1].update(generator_names)
            for condition in generator.ifs:
                self.visit(condition)
        visit_value()
        self._pop_scope()

    def visit_ListComp(self, node: ast.ListComp) -> None:  # noqa: N802
        self._visit_comprehension(node.generators, lambda: self.visit(node.elt))

    def visit_SetComp(self, node: ast.SetComp) -> None:  # noqa: N802
        self._visit_comprehension(node.generators, lambda: self.visit(node.elt))

    def visit_GeneratorExp(self, node: ast.GeneratorExp) -> None:  # noqa: N802
        self._visit_comprehension(node.generators, lambda: self.visit(node.elt))

    def visit_DictComp(self, node: ast.DictComp) -> None:  # noqa: N802
        self._visit_comprehension(node.generators, lambda: (self.visit(node.key), self.visit(node.value)))

    def visit_Name(self, node: ast.Name) -> None:  # noqa: N802
        if isinstance(node.ctx, ast.Load) and not self._is_defined(node.id):
            self._report_undefined(node)


def validate_python_config_content(path: str | Path, content: str) -> list[ConfigValidationIssue]:
    target = Path(path)
    try:
        tree = ast.parse(content, filename=str(target))
    except SyntaxError as exc:
        return [
            ConfigValidationIssue(
                path=target,
                message=exc.msg or "Python 语法错误",
                line=exc.lineno,
                column=exc.offset,
                code="syntax-error",
            )
        ]

    validator = _UndefinedNameValidator(target)
    validator.visit(tree)
    validator.issues.extend(_validate_runtime_contract(target, tree))
    return sorted(
        validator.issues,
        key=lambda item: (
            str(item.path),
            -1 if item.line is None else item.line,
            -1 if item.column is None else item.column,
            item.message,
        ),
    )


def validate_runtime_definition_files() -> list[ConfigValidationIssue]:
    issues: list[ConfigValidationIssue] = []
    for directory in [KPI_SPECS_DIR, DERIVED_SIGNALS_DIR]:
        for path in _iter_python_files(directory):
            if _is_system_template_path(path):
                continue
            issues.extend(validate_python_config_content(path, path.read_text(encoding="utf-8")))
    issues.extend(_collect_runtime_dependency_issues())
    return sorted(
        issues,
        key=lambda item: (
            str(item.path),
            -1 if item.line is None else item.line,
            -1 if item.column is None else item.column,
            item.message,
        ),
    )


def _collect_runtime_dependency_issues() -> list[ConfigValidationIssue]:
    issues: dict[tuple[str, str], ConfigValidationIssue] = {}
    kpi_definitions = load_kpi_definitions()
    derived_definitions = load_derived_signal_definitions()
    derived_lookup = {
        str(definition.get("name", "")).strip(): definition
        for definition in derived_definitions
        if str(definition.get("name", "")).strip()
    }
    visiting: set[str] = set()
    validated: set[str] = set()

    def add_issue(path: Path | str, message: str, code: str) -> None:
        issue_path = Path(path) if path else CONFIG_DIR
        key = (str(issue_path), message)
        issues[key] = ConfigValidationIssue(path=issue_path, message=message, code=code)

    def validate_derived(signal_name: str, owner_path: Path | str, owner_name: str, owner_kind: str) -> bool:
        normalized_name = str(signal_name).strip()
        if not normalized_name:
            return True
        if normalized_name in validated:
            return True
        if normalized_name in visiting:
            add_issue(owner_path, f"{owner_kind} {owner_name} 的 derived_inputs 存在循环依赖: {normalized_name}", "derived-cycle")
            return False
        definition = derived_lookup.get(normalized_name)
        if definition is None:
            add_issue(owner_path, f"{owner_kind} {owner_name} 的 derived_inputs 引用了不存在的派生量: {normalized_name}", "missing-derived-input")
            return False
        visiting.add(normalized_name)
        valid = True
        for dependency_name in definition.get("derived_inputs", []):
            if not validate_derived(
                str(dependency_name).strip(),
                Path(str(definition.get("module_path", ""))) if definition.get("module_path") else owner_path,
                normalized_name,
                "派生量",
            ):
                valid = False
        visiting.remove(normalized_name)
        if valid:
            validated.add(normalized_name)
        return valid

    for definition in derived_definitions:
        signal_name = str(definition.get("name", "")).strip()
        if not signal_name:
            continue
        for dependency_name in definition.get("derived_inputs", []):
            validate_derived(
                str(dependency_name).strip(),
                Path(str(definition.get("module_path", ""))) if definition.get("module_path") else CONFIG_DIR,
                signal_name,
                "派生量",
            )

    for definition in kpi_definitions:
        kpi_name = str(definition.get("name", "")).strip() or "unnamed_kpi"
        for dependency_name in definition.get("derived_inputs", []):
            validate_derived(
                str(dependency_name).strip(),
                Path(str(definition.get("module_path", ""))) if definition.get("module_path") else CONFIG_DIR,
                kpi_name,
                "KPI",
            )

    return sorted(
        issues.values(),
        key=lambda item: (
            str(item.path),
            -1 if item.line is None else item.line,
            -1 if item.column is None else item.column,
            item.message,
        ),
    )


def _extract_literal_dict(node: ast.AST) -> dict[str, ast.AST]:
    if not isinstance(node, ast.Dict):
        return {}
    values: dict[str, ast.AST] = {}
    for key_node, value_node in zip(node.keys, node.values):
        if isinstance(key_node, ast.Constant) and isinstance(key_node.value, str):
            values[str(key_node.value)] = value_node
    return values


def _extract_string_constant(node: ast.AST | None) -> str | None:
    if isinstance(node, ast.Constant) and isinstance(node.value, str):
        return str(node.value)
    return None


def _validate_runtime_contract(path: Path, tree: ast.AST) -> list[ConfigValidationIssue]:
    if _is_system_template_path(path):
        return []

    issues: list[ConfigValidationIssue] = []
    function_names = {
        node.name
        for node in ast.walk(tree)
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef))
    }
    definition_node: ast.AST | None = None
    for node in tree.body if isinstance(tree, ast.Module) else []:
        if isinstance(node, ast.Assign):
            for target in node.targets:
                if isinstance(target, ast.Name) and target.id == "KPI_DEFINITION":
                    definition_node = node.value
                    break
        if definition_node is not None:
            break
    definition_map = _extract_literal_dict(definition_node) if definition_node is not None else {}
    if not definition_map and path.parent != KPI_SPECS_DIR:
        return []
    kpi_name = _extract_string_constant(definition_map.get("name"))
    trend_source = _extract_string_constant(definition_map.get("trend_source"))

    if "calculate_kpi_series" not in function_names:
        issues.append(
            ConfigValidationIssue(
                path=path,
                message="KPI 必须定义 calculate_kpi_series(dataframe)，用于展示单值结果的连续计算过程",
                code="kpi-missing-series",
            )
        )
    if not trend_source:
        issues.append(
            ConfigValidationIssue(
                path=path,
                message="KPI_DEFINITION 必须提供 trend_source，且该字段必须与 name 完全一致",
                code="kpi-missing-trend-source",
            )
        )
    elif kpi_name and trend_source != kpi_name:
        issues.append(
            ConfigValidationIssue(
                path=path,
                message=f'trend_source 必须与 KPI name 完全一致: {kpi_name}',
                code="kpi-trend-source-mismatch",
            )
        )
    return issues


def _is_system_template_path(path: Path) -> bool:
    return path.name == "00_example_and_guide.py"


def _ensure_supporting_files() -> None:
    RULE_SPECS_DIR.mkdir(parents=True, exist_ok=True)
    KPI_SPECS_DIR.mkdir(parents=True, exist_ok=True)
    DERIVED_SIGNALS_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    if not FORMULA_SIGNALS_PATH.exists():
        FORMULA_SIGNALS_PATH.write_text(json.dumps({"signals": []}, ensure_ascii=False, indent=2), encoding="utf-8")
    if not KPI_GROUPS_PATH.exists():
        KPI_GROUPS_PATH.write_text(json.dumps({"groups": []}, ensure_ascii=False, indent=2), encoding="utf-8")
    if not CHART_VIEW_STATE_PATH.exists():
        CHART_VIEW_STATE_PATH.write_text(
            json.dumps({"active_sheet": 0, "sheets": [{"name": "工作表 1", "panels": []}]}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    guide_template_path = REPORT_TEMPLATES_DIR / "00_example_and_guide.html"
    if not guide_template_path.exists():
        guide_template_path.write_text(DEFAULT_REPORT_GUIDE_TEMPLATE, encoding="utf-8")
    demo_template_path = REPORT_TEMPLATES_DIR / "demo_template.html"
    if not demo_template_path.exists():
        demo_template_path.write_text(DEFAULT_REPORT_TEMPLATE, encoding="utf-8")
    _refresh_generated_guide_files()


def _guide_iter_python_files(directory: Path) -> list[Path]:
    if not directory.exists():
        return []
    return sorted(path for path in directory.glob("*.py") if path.is_file() and path.name != "__init__.py")


def _load_derived_signal_catalog_for_guides() -> list[dict[str, str]]:
    catalog: list[dict[str, str]] = []
    for path in _guide_iter_python_files(DERIVED_SIGNALS_DIR):
        module = _load_python_module(path, "derived_signal_guide")
        if _is_system_template_path(path):
            continue
        definition = getattr(module, "DERIVED_SIGNAL_DEFINITION", {}) or {}
        name = str(definition.get("name", "")).strip()
        if not name:
            continue
        catalog.append(
            {
                "name": name,
                "title": str(definition.get("title", "")).strip(),
                "description": str(definition.get("description", "")).strip(),
                "algorithm_summary": str(definition.get("algorithm_summary", "")).strip(),
            }
        )
    return sorted(catalog, key=lambda item: item["name"].lower())


def _render_derived_catalog_for_kpi_guide() -> str:
    catalog = _load_derived_signal_catalog_for_guides()
    if not catalog:
        return "- 当前尚未定义可复用派生量。"

    lines: list[str] = []
    for index, item in enumerate(catalog, start=1):
        summary = item["algorithm_summary"] or item["description"] or "未填写"
        title = item["title"] or "未填写title"
        lines.append(f"{index}. {item['name']} - {title} - {summary}")
    return "\n".join(lines).rstrip()


def _render_derived_catalog_for_derived_guide() -> str:
    catalog = _load_derived_signal_catalog_for_guides()
    if not catalog:
        return "- 当前尚未定义可复用派生量。"

    lines: list[str] = []
    for index, item in enumerate(catalog, start=1):
        summary = item["algorithm_summary"] or item["description"] or "未填写"
        title = item["title"] or "未填写title"
        lines.append(f"{index}. {item['name']} - {title} - {summary}")
    return "\n".join(lines).rstrip()


def _ordered_raw_input_names(raw_names: list[str] | set[str] | tuple[str, ...]) -> list[str]:
    normalized = sorted({str(item).strip() for item in raw_names if str(item).strip()}, key=str.lower)
    if "time_s" in normalized:
        normalized.remove("time_s")
        normalized.insert(0, "time_s")
    return normalized


def _render_raw_input_catalog_for_guides() -> str:
    required_by: dict[str, list[str]] = {"time_s": []}
    for directory, definition_name, owner_prefix in [
        (KPI_SPECS_DIR, "KPI_DEFINITION", "KPI"),
        (DERIVED_SIGNALS_DIR, "DERIVED_SIGNAL_DEFINITION", "派生量"),
    ]:
        for path in _guide_iter_python_files(directory):
            if _is_system_template_path(path):
                continue
            module = _load_python_module(path, f"guide_raw_inputs_{directory.name}")
            definition = getattr(module, definition_name, {}) or {}
            owner_name = str(definition.get("name", "")).strip()
            if not owner_name:
                continue
            for signal_name in definition.get("raw_inputs", []):
                normalized_signal = str(signal_name).strip()
                if not normalized_signal:
                    continue
                required_by.setdefault(normalized_signal, []).append(f"{owner_prefix}:{owner_name}")
    if not required_by:
        return "- 当前尚未定义标准输入量。"

    lines: list[str] = []
    for index, signal_name in enumerate(_ordered_raw_input_names(list(required_by)), start=1):
        description = RAW_INPUT_DESCRIPTIONS.get(signal_name, "请补充该输入量的物理意义与单位说明。")
        lines.append(f"{index}. {signal_name} - {description}")
    return "\n".join(lines).rstrip()


def _render_kpi_guide_template() -> str:
    return (
        KPI_GUIDE_TEMPLATE.replace("{{DERIVED_CATALOG}}", _render_derived_catalog_for_kpi_guide())
        .replace("{{RAW_INPUT_CATALOG}}", _render_raw_input_catalog_for_guides())
    )


def _refresh_generated_guide_files() -> None:
    (DERIVED_SIGNALS_DIR / "00_example_and_guide.py").write_text(
        DEFAULT_DERIVED_SIGNAL_GUIDE_TEMPLATE.replace("{{DERIVED_CATALOG}}", _render_derived_catalog_for_derived_guide()).replace(
            "{{RAW_INPUT_CATALOG}}", _render_raw_input_catalog_for_guides()
        ),
        encoding="utf-8",
    )
    (KPI_SPECS_DIR / "00_example_and_guide.py").write_text(_render_kpi_guide_template(), encoding="utf-8")


def _format_required_by_owner(owner: str) -> str:
    normalized = str(owner).strip()
    if normalized.startswith("KPI:"):
        return f"KPI: {normalized.split(':', 1)[1]}"
    if normalized.startswith("DERIVED:"):
        return f"派生量: {normalized.split(':', 1)[1]}"
    return normalized


def _compose_display_name(name: str, title: str, fallback: str) -> str:
    normalized_name = str(name or "").strip()
    normalized_title = str(title or "").strip()
    if normalized_name and normalized_title:
        return f"{normalized_name} {normalized_title}"
    if normalized_title:
        return normalized_title
    if normalized_name:
        return normalized_name
    return fallback


def _definition_display_name(module, definition_name: str, fallback: str) -> str:  # noqa: ANN001
    definition = getattr(module, definition_name, {}) or {}
    return _compose_display_name(str(definition.get("name", "")), str(definition.get("title", "")), fallback)


def _deprecated_payload(kind: str) -> dict[str, Any]:
    return {
        "deprecated": True,
        "kind": kind,
        "note": f"该文件已废弃，仅为兼容保留。当前生效入口请改用 src/tcs_smart_analyzer/config/{'rule_specs/' if kind == 'rules' else 'kpi_specs/'} 目录。",
        "current_entry": "rule_specs/" if kind == "rules" else "kpi_specs/",
        "read_only_reference": True,
        "definitions": [],
    }


def sync_legacy_deprecated_configs() -> None:
    for path, payload in [
        (LEGACY_RULE_DEFINITIONS_PATH, _deprecated_payload("rules")),
        (LEGACY_KPI_DEFINITIONS_PATH, _deprecated_payload("kpis")),
    ]:
        if path.exists():
            os.chmod(path, 0o666)
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        os.chmod(path, 0o444)


def _read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _load_python_module(module_path: Path, namespace: str):
    module_name = f"tcs_smart_analyzer_dynamic_{namespace}_{module_path.stem}"
    spec = importlib.util.spec_from_file_location(module_name, module_path)
    if spec is None or spec.loader is None:
        raise ImportError(f"无法加载配置模块: {module_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _iter_python_files(directory: Path) -> list[Path]:
    _ensure_supporting_files()
    if not directory.exists():
        return []
    return sorted(path for path in directory.glob("*.py") if path.is_file() and path.name != "__init__.py")


def _extract_definition_name_for_directory(directory: Path, content: str) -> str:
    if directory == KPI_SPECS_DIR:
        return extract_kpi_name_from_text(content)
    if directory == DERIVED_SIGNALS_DIR:
        return extract_derived_signal_name_from_text(content)
    return ""


def _normalize_definition_file_names(directory: Path) -> list[Path]:
    if directory not in {KPI_SPECS_DIR, DERIVED_SIGNALS_DIR} or not directory.exists():
        return []
    renamed_paths: list[Path] = []
    for path in sorted(directory.glob("*.py")):
        if not path.is_file() or path.name == "__init__.py" or _is_system_template_path(path):
            continue
        try:
            definition_name = _extract_definition_name_for_directory(directory, path.read_text(encoding="utf-8"))
        except OSError:
            continue
        if not definition_name:
            continue
        desired_stem = _slugify_file_stem(definition_name, path.stem)
        desired_path = path.with_name(f"{desired_stem}{path.suffix}")
        if desired_path == path or desired_path.exists():
            continue
        path.rename(desired_path)
        renamed_paths.append(desired_path)
    return renamed_paths


def _path_signature(path: Path, pattern: str | None = None) -> tuple[object, ...]:
    if not path.exists():
        return (str(path), "missing")
    if path.is_file():
        stat = path.stat()
        return (str(path), stat.st_mtime_ns, stat.st_size)
    entries: list[tuple[object, ...]] = []
    for item in sorted(path.glob(pattern or "*")):
        if not item.is_file():
            continue
        stat = item.stat()
        entries.append((item.name, stat.st_mtime_ns, stat.st_size))
    return tuple(entries)


def _slugify_file_stem(name: str, fallback: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9_]+", "_", name.strip()).strip("_").lower()
    return slug or fallback


def _default_kpi_group() -> dict[str, Any]:
    return {
        "key": "__all_kpis__",
        "name": "默认组（全部 KPI）",
        "kpis": [],
        "is_builtin": True,
    }


def load_kpi_groups() -> list[dict[str, Any]]:
    _ensure_supporting_files()
    raw = _read_json(KPI_GROUPS_PATH)
    groups = raw.get("groups", []) if isinstance(raw, dict) else []
    default_group = _default_kpi_group()
    normalized: list[dict[str, Any]] = [default_group]
    for item in groups:
        if not isinstance(item, dict):
            continue
        key = str(item.get("key", "")).strip()
        name = str(item.get("name", "")).strip()
        kpis = [str(value).strip() for value in item.get("kpis", []) if str(value).strip()]
        kpis = list(dict.fromkeys(kpis))
        if not key:
            continue
        if key == "__all_kpis__":
            default_group["kpis"] = kpis
            continue
        if not name:
            continue
        normalized.append({"key": key, "name": name, "kpis": kpis, "is_builtin": False})
    return normalized


def save_kpi_group(name: str, kpis: list[str], key: str | None = None) -> Path:
    _ensure_supporting_files()
    normalized_name = name.strip()
    if not normalized_name:
        raise ValueError("KPI 分组名称不能为空")
    normalized_key = (key or _slugify_file_stem(normalized_name, "kpi_group")).strip()
    payload = [group for group in load_kpi_groups() if not group.get("is_builtin")]
    payload = [group for group in payload if str(group.get("key")) != normalized_key]
    if normalized_key == "__all_kpis__":
        payload.insert(
            0,
            {
                "key": normalized_key,
                "name": _default_kpi_group()["name"],
                "kpis": list(dict.fromkeys(str(item).strip() for item in kpis if str(item).strip())),
            },
        )
        KPI_GROUPS_PATH.write_text(json.dumps({"groups": payload}, ensure_ascii=False, indent=2), encoding="utf-8")
        return KPI_GROUPS_PATH
    payload.append(
        {
            "key": normalized_key,
            "name": normalized_name,
            "kpis": list(dict.fromkeys(str(item).strip() for item in kpis if str(item).strip())),
        }
    )
    KPI_GROUPS_PATH.write_text(json.dumps({"groups": payload}, ensure_ascii=False, indent=2), encoding="utf-8")
    return KPI_GROUPS_PATH


def delete_kpi_group(key: str) -> Path:
    _ensure_supporting_files()
    normalized_key = key.strip()
    if normalized_key == "__all_kpis__":
        raise ValueError("默认组不能删除")
    payload = [group for group in load_kpi_groups() if not group.get("is_builtin") and str(group.get("key")) != normalized_key]
    KPI_GROUPS_PATH.write_text(json.dumps({"groups": payload}, ensure_ascii=False, indent=2), encoding="utf-8")
    return KPI_GROUPS_PATH


def _resolve_group_filter(group_key: str | None) -> set[str] | None:
    if not group_key or group_key == "__all_kpis__":
        return None
    for group in load_kpi_groups():
        if str(group.get("key")) == group_key:
            return {str(item).strip() for item in group.get("kpis", []) if str(item).strip()}
    return set()


def get_kpi_group_order(group_key: str | None) -> list[str] | None:
    if not group_key:
        group_key = "__all_kpis__"
    for group in load_kpi_groups():
        if str(group.get("key")) == group_key:
            return [str(item).strip() for item in group.get("kpis", []) if str(item).strip()]
    return None if group_key == "__all_kpis__" else []


def get_interface_mapping_actual_name_column_count() -> int:
    if not INTERFACE_MAPPING_PATH.exists():
        return DEFAULT_INTERFACE_ACTUAL_NAME_COLUMN_COUNT
    workbook = load_workbook(INTERFACE_MAPPING_PATH)
    if INTERFACE_MAPPING_METADATA_SHEET in workbook.sheetnames:
        metadata_sheet = workbook[INTERFACE_MAPPING_METADATA_SHEET]
        for key, value, *_ in metadata_sheet.iter_rows(min_row=1, values_only=True):
            if str(key or "").strip() == "actual_signal_name_column_count":
                try:
                    return max(DEFAULT_INTERFACE_ACTUAL_NAME_COLUMN_COUNT, int(value or DEFAULT_INTERFACE_ACTUAL_NAME_COLUMN_COUNT))
                except (TypeError, ValueError):
                    break
    max_count = DEFAULT_INTERFACE_ACTUAL_NAME_COLUMN_COUNT
    for sheet_name in [SYSTEM_MAPPING_SHEET, CUSTOM_MAPPING_SHEET]:
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        header_row = [str(cell or "").strip().lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])]
        actual_count = sum(1 for cell in header_row if cell.startswith("actual_signal_name_"))
        if actual_count > DEFAULT_INTERFACE_ACTUAL_NAME_COLUMN_COUNT:
            extra_column_start = (2 if sheet_name == SYSTEM_MAPPING_SHEET else 1) + DEFAULT_INTERFACE_ACTUAL_NAME_COLUMN_COUNT
            has_legacy_extra_values = any(
                str(cell or "").strip()
                for row in sheet.iter_rows(min_row=2, values_only=True)
                for cell in row[extra_column_start:]
            )
            if not has_legacy_extra_values:
                actual_count = DEFAULT_INTERFACE_ACTUAL_NAME_COLUMN_COUNT
        max_count = max(max_count, actual_count)
    return max_count


def collect_kpi_raw_input_requirements(group_key: str | None = None) -> dict[str, list[str]]:
    required_by: dict[str, list[str]] = {}
    for definition in load_kpi_definitions(group_key):
        owner = f"KPI:{definition.get('name', 'kpi')}"
        for signal_name in definition.get("raw_inputs", []):
            signal_key = str(signal_name).strip()
            if signal_key:
                required_by.setdefault(signal_key, []).append(owner)
    return required_by


def list_kpi_raw_input_signals(group_key: str | None = None) -> list[str]:
    return sorted(collect_kpi_raw_input_requirements(group_key))


def collect_derived_signal_raw_input_requirements(signal_names: set[str] | None = None) -> dict[str, list[str]]:
    required_by: dict[str, list[str]] = {}
    for definition in load_derived_signal_definitions():
        signal_name = str(definition.get("name", "")).strip()
        if signal_names is not None and signal_name not in signal_names:
            continue
        owner = f"DERIVED:{signal_name or 'derived_signal'}"
        for raw_name in definition.get("raw_inputs", []):
            signal_key = str(raw_name).strip()
            if signal_key:
                required_by.setdefault(signal_key, []).append(owner)
    return required_by


def _resolve_required_derived_signal_names_from_definitions(
    kpi_definitions: list[dict[str, Any]],
    derived_definitions: list[dict[str, Any]],
    *,
    strict: bool = True,
) -> set[str]:
    lookup = {
        str(item.get("name", "")).strip(): item
        for item in derived_definitions
        if str(item.get("name", "")).strip()
    }
    resolved: set[str] = set()
    visiting: set[str] = set()

    def visit(signal_name: str) -> bool:
        if not signal_name or signal_name in resolved:
            return True
        if signal_name in visiting:
            if strict:
                raise ValueError(f"检测到派生量循环依赖: {signal_name}")
            return False
        definition = lookup.get(signal_name)
        if definition is None:
            if strict:
                raise ValueError(f"KPI 依赖了不存在的派生量: {signal_name}")
            return False
        visiting.add(signal_name)
        valid = True
        for dependency_name in definition.get("derived_inputs", []):
            if not visit(str(dependency_name).strip()):
                valid = False
        visiting.remove(signal_name)
        if not valid and not strict:
            return False
        resolved.add(signal_name)
        return True

    for definition in kpi_definitions:
        for signal_name in definition.get("derived_inputs", []):
            visit(str(signal_name).strip())
    return resolved


def resolve_required_derived_signal_names(group_key: str | None = None, *, strict: bool = True) -> set[str]:
    return _resolve_required_derived_signal_names_from_definitions(
        load_kpi_definitions(group_key),
        load_derived_signal_definitions(),
        strict=strict,
    )


def list_required_raw_input_signals(group_key: str | None = None) -> list[str]:
    required_by = collect_kpi_raw_input_requirements(group_key)
    for signal_name, owners in collect_derived_signal_raw_input_requirements(resolve_required_derived_signal_names(group_key, strict=False)).items():
        required_by.setdefault(signal_name, []).extend(owners)
    required_by.setdefault("time_s", [])
    return _ordered_raw_input_names(list(required_by))


def list_rule_spec_entries() -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    for path in _iter_python_files(RULE_SPECS_DIR):
        module = _load_python_module(path, "rule_list")
        entries.append(
            {
                "path": path,
                "display_name": str(getattr(module, "DISPLAY_NAME", path.stem)),
                "is_template": bool(getattr(module, "IS_TEMPLATE", False)),
            }
        )
    return sorted(entries, key=lambda item: (not item["is_template"], str(item["display_name"])))


def list_kpi_spec_entries() -> list[dict[str, Any]]:
    _normalize_definition_file_names(KPI_SPECS_DIR)
    entries: list[dict[str, Any]] = []
    for path in _iter_python_files(KPI_SPECS_DIR):
        module = _load_python_module(path, "kpi_list")
        is_template = _is_system_template_path(path)
        entries.append(
            {
                "path": path,
                "display_name": "示例与详细讲解" if is_template else _definition_display_name(module, "KPI_DEFINITION", path.stem),
                "is_template": is_template,
            }
        )
    return sorted(entries, key=lambda item: (not item["is_template"], str(item["display_name"])))


def list_derived_signal_spec_entries() -> list[dict[str, Any]]:
    _normalize_definition_file_names(DERIVED_SIGNALS_DIR)
    entries: list[dict[str, Any]] = []
    for path in _iter_python_files(DERIVED_SIGNALS_DIR):
        module = _load_python_module(path, "derived_signal_list")
        is_template = _is_system_template_path(path)
        entries.append(
            {
                "path": path,
                "display_name": "示例与详细讲解" if is_template else _definition_display_name(module, "DERIVED_SIGNAL_DEFINITION", path.stem),
                "is_template": is_template,
            }
        )
    return sorted(entries, key=lambda item: (not item["is_template"], str(item["display_name"])))


def list_report_template_entries() -> list[dict[str, Any]]:
    _ensure_supporting_files()
    entries: list[dict[str, Any]] = []
    for path in sorted(REPORT_TEMPLATES_DIR.glob("*.html")):
        entries.append(
            {
                "path": path,
                "display_name": "示例与详细讲解" if path.name == "00_example_and_guide.html" else path.stem,
                "is_template": path.name == "00_example_and_guide.html",
            }
        )
    return sorted(entries, key=lambda item: (not item["is_template"], str(item["display_name"])))


def load_rule_definitions() -> list[dict[str, Any]]:
    return _load_rule_definitions_cached(_path_signature(RULE_SPECS_DIR, "*.py"))


@lru_cache(maxsize=8)
def _load_rule_definitions_cached(_signature: tuple[object, ...]) -> list[dict[str, Any]]:
    definitions: list[dict[str, Any]] = []
    for path in _iter_python_files(RULE_SPECS_DIR):
        module = _load_python_module(path, "rule_defs")
        if bool(getattr(module, "IS_TEMPLATE", False)):
            continue
        definition = deepcopy(getattr(module, "RULE_DEFINITION", {}))
        if not definition:
            continue
        definition["module_path"] = str(path)
        definitions.append(definition)
    return definitions


def load_rule_plugins() -> list[dict[str, Any]]:
    return _load_rule_plugins_cached(_path_signature(RULE_SPECS_DIR, "*.py"))


@lru_cache(maxsize=8)
def _load_rule_plugins_cached(_signature: tuple[object, ...]) -> list[dict[str, Any]]:
    plugins: list[dict[str, Any]] = []
    for path in _iter_python_files(RULE_SPECS_DIR):
        module = _load_python_module(path, "rule_plugins")
        if bool(getattr(module, "IS_TEMPLATE", False)):
            continue
        definition = deepcopy(getattr(module, "RULE_DEFINITION", {}))
        evaluate_rule = getattr(module, "evaluate_rule", None)
        if not definition or evaluate_rule is None:
            continue
        definition["module_path"] = str(path)
        plugins.append({"definition": definition, "evaluate_rule": evaluate_rule, "path": path})
    return plugins


def load_kpi_definitions(group_key: str | None = None) -> list[dict[str, Any]]:
    _normalize_definition_file_names(KPI_SPECS_DIR)
    return _load_kpi_definitions_cached(_path_signature(KPI_SPECS_DIR, "*.py"), _path_signature(KPI_GROUPS_PATH), group_key)


@lru_cache(maxsize=32)
def _load_kpi_definitions_cached(
    _spec_signature: tuple[object, ...],
    _group_signature: tuple[object, ...],
    group_key: str | None,
) -> list[dict[str, Any]]:
    definitions: list[dict[str, Any]] = []
    allowed_kpis = _resolve_group_filter(group_key)
    group_order = get_kpi_group_order(group_key)
    for path in _iter_python_files(KPI_SPECS_DIR):
        module = _load_python_module(path, "kpi_defs")
        if _is_system_template_path(path):
            continue
        definition = deepcopy(getattr(module, "KPI_DEFINITION", {}))
        if not definition:
            continue
        if allowed_kpis is not None and str(definition.get("name", "")).strip() not in allowed_kpis:
            continue
        definition["module_path"] = str(path)
        definition["display_name"] = _compose_display_name(definition.get("name", ""), definition.get("title", ""), path.stem)
        definitions.append(definition)
    if group_order is not None:
        order_index = {name: index for index, name in enumerate(group_order)}
        definitions.sort(key=lambda item: (order_index.get(str(item.get("name", "")), len(order_index)), str(item.get("name", ""))))
    return definitions


def load_kpi_plugins(group_key: str | None = None) -> list[dict[str, Any]]:
    _normalize_definition_file_names(KPI_SPECS_DIR)
    return _load_kpi_plugins_cached(_path_signature(KPI_SPECS_DIR, "*.py"), _path_signature(KPI_GROUPS_PATH), group_key)


@lru_cache(maxsize=32)
def _load_kpi_plugins_cached(
    _spec_signature: tuple[object, ...],
    _group_signature: tuple[object, ...],
    group_key: str | None,
) -> list[dict[str, Any]]:
    plugins: list[dict[str, Any]] = []
    allowed_kpis = _resolve_group_filter(group_key)
    group_order = get_kpi_group_order(group_key)
    for path in _iter_python_files(KPI_SPECS_DIR):
        module = _load_python_module(path, "kpi_plugins")
        if _is_system_template_path(path):
            continue
        definition = deepcopy(getattr(module, "KPI_DEFINITION", {}))
        calculate_kpi = getattr(module, "calculate_kpi", None)
        calculate_kpi_series = getattr(module, "calculate_kpi_series", None)
        if not definition or calculate_kpi is None:
            continue
        if allowed_kpis is not None and str(definition.get("name", "")).strip() not in allowed_kpis:
            continue
        definition["module_path"] = str(path)
        definition["display_name"] = _compose_display_name(definition.get("name", ""), definition.get("title", ""), path.stem)
        plugins.append(
            {
                "definition": definition,
                "calculate_kpi": calculate_kpi,
                "calculate_kpi_series": calculate_kpi_series,
                "path": path,
            }
        )
    if group_order is not None:
        order_index = {name: index for index, name in enumerate(group_order)}
        plugins.sort(key=lambda item: (order_index.get(str(item.get("definition", {}).get("name", "")), len(order_index)), str(item.get("definition", {}).get("name", ""))))
    return plugins


def load_derived_signal_definitions() -> list[dict[str, Any]]:
    _normalize_definition_file_names(DERIVED_SIGNALS_DIR)
    return _load_derived_signal_definitions_cached(_path_signature(DERIVED_SIGNALS_DIR, "*.py"))


@lru_cache(maxsize=8)
def _load_derived_signal_definitions_cached(_signature: tuple[object, ...]) -> list[dict[str, Any]]:
    definitions: list[dict[str, Any]] = []
    for path in _iter_python_files(DERIVED_SIGNALS_DIR):
        module = _load_python_module(path, "derived_signal_defs")
        if _is_system_template_path(path):
            continue
        definition = deepcopy(getattr(module, "DERIVED_SIGNAL_DEFINITION", {}))
        if not definition:
            continue
        definition["module_path"] = str(path)
        definition["display_name"] = _compose_display_name(definition.get("name", ""), definition.get("title", ""), path.stem)
        definitions.append(definition)
    return definitions


def load_derived_signal_plugins() -> list[dict[str, Any]]:
    _normalize_definition_file_names(DERIVED_SIGNALS_DIR)
    return _load_derived_signal_plugins_cached(_path_signature(DERIVED_SIGNALS_DIR, "*.py"))


@lru_cache(maxsize=8)
def _load_derived_signal_plugins_cached(_signature: tuple[object, ...]) -> list[dict[str, Any]]:
    plugins: list[dict[str, Any]] = []
    for path in _iter_python_files(DERIVED_SIGNALS_DIR):
        module = _load_python_module(path, "derived_signal_plugins")
        if _is_system_template_path(path):
            continue
        definition = deepcopy(getattr(module, "DERIVED_SIGNAL_DEFINITION", {}))
        calculate_signal = getattr(module, "calculate_signal", None)
        if not definition or calculate_signal is None:
            continue
        definition["module_path"] = str(path)
        definition["display_name"] = _compose_display_name(definition.get("name", ""), definition.get("title", ""), path.stem)
        plugins.append({"definition": definition, "calculate_signal": calculate_signal, "path": path})
    return plugins


def build_default_rule_settings() -> dict[str, dict[str, Any]]:
    settings: dict[str, dict[str, Any]] = {}
    for definition in load_kpi_definitions():
        kpi_name = str(definition.get("name", "")).strip()
        if not kpi_name:
            continue
        settings[kpi_name] = {
            "threshold": definition.get("threshold"),
            "source": definition.get("source", "kpi_definition"),
            "enabled": definition.get("enabled", True),
            "unit": definition.get("unit", ""),
        }
    return settings


def get_config_file_paths() -> dict[str, Path]:
    _ensure_supporting_files()
    sync_legacy_deprecated_configs()
    return {
        "rule_specs_dir": RULE_SPECS_DIR,
        "kpi_specs_dir": KPI_SPECS_DIR,
        "derived_signals_dir": DERIVED_SIGNALS_DIR,
        "kpi_groups": KPI_GROUPS_PATH,
        "interface_mapping": INTERFACE_MAPPING_PATH,
        "report_templates_dir": REPORT_TEMPLATES_DIR,
        "formula_signals": FORMULA_SIGNALS_PATH,
        "chart_view_state": CHART_VIEW_STATE_PATH,
        "legacy_rule_definitions": LEGACY_RULE_DEFINITIONS_PATH,
        "legacy_kpi_definitions": LEGACY_KPI_DEFINITIONS_PATH,
    }


def read_text_config_file(path: str | Path) -> str:
    return Path(path).read_text(encoding="utf-8")


def write_text_config_file(path: str | Path, content: str) -> Path:
    target = Path(path)
    target.write_text(content, encoding="utf-8")
    return target


def align_python_config_file_name(path: str | Path, definition_name: str) -> Path:
    target = Path(path)
    normalized_name = str(definition_name or "").strip()
    if not normalized_name:
        return target
    desired_stem = _slugify_file_stem(normalized_name, target.stem)
    desired_path = target.with_name(f"{desired_stem}{target.suffix}")
    if desired_path == target:
        return target
    if desired_path.exists():
        raise FileExistsError(f"目标文件已存在: {desired_path.name}")
    target.rename(desired_path)
    return desired_path


def create_rule_spec_file(display_name: str, rule_id: str | None = None) -> Path:
    _ensure_supporting_files()
    safe_stem = _slugify_file_stem(display_name, "new_rule")
    target = RULE_SPECS_DIR / f"{safe_stem}.py"
    if target.exists():
        raise FileExistsError(f"规则文件已存在: {target.name}")
    final_rule_id = rule_id or safe_stem.upper().replace("_", "-")
    target.write_text(DEFAULT_RULE_TEMPLATE.format(display_name=display_name, rule_id=final_rule_id), encoding="utf-8")
    return target


def create_kpi_spec_file(display_name: str, kpi_name: str | None = None) -> Path:
    _ensure_supporting_files()
    safe_stem = _slugify_file_stem(display_name, "new_kpi")
    target = KPI_SPECS_DIR / f"{safe_stem}.py"
    if target.exists():
        raise FileExistsError(f"KPI 文件已存在: {target.name}")
    final_kpi_name = kpi_name or safe_stem
    target.write_text(DEFAULT_KPI_TEMPLATE.format(kpi_name=final_kpi_name, title=display_name), encoding="utf-8")
    return target


def _next_available_config_path(directory: Path, stem: str, suffix: str) -> Path:
    candidate = directory / f"{stem}{suffix}"
    if not candidate.exists():
        return candidate
    index = 1
    while True:
        candidate = directory / f"{stem}_{index}{suffix}"
        if not candidate.exists():
            return candidate
        index += 1


def _build_guide_copy_content(guide_path: Path, name: str, title: str) -> str:
    content = guide_path.read_text(encoding="utf-8")
    content = re.sub(r'"name":\s*"[^"]*"', f'"name": "{name}"', content, count=1)
    return re.sub(r'"title":\s*"[^"]*"', f'"title": "{title}"', content, count=1)


def _extract_definition_field_value(content: str, definition_name: str, field_name: str) -> str:
    match = re.search(
        rf"{re.escape(definition_name)}\s*=\s*\{{.*?\"{re.escape(field_name)}\"\s*:\s*\"([^\"]*)\"",
        content,
        flags=re.DOTALL,
    )
    return "" if match is None else str(match.group(1)).strip()


def extract_kpi_name_from_text(content: str) -> str:
    return _extract_definition_field_value(content, "KPI_DEFINITION", "name")


def extract_derived_signal_name_from_text(content: str) -> str:
    return _extract_definition_field_value(content, "DERIVED_SIGNAL_DEFINITION", "name")


def _rename_identifier_references_in_python_config(content: str, old_name: str, new_name: str) -> str:
    if not old_name or not new_name or old_name == new_name:
        return content

    old_token = f'"{old_name}"'
    new_token = f'"{new_name}"'
    updated_lines: list[str] = []
    inside_derived_inputs = False
    bracket_balance = 0

    for line in content.splitlines(keepends=True):
        updated_line = line

        if inside_derived_inputs or '"derived_inputs"' in updated_line:
            updated_line = updated_line.replace(old_token, new_token)
            bracket_balance += updated_line.count("[") - updated_line.count("]")
            inside_derived_inputs = bracket_balance > 0
            if not inside_derived_inputs:
                bracket_balance = 0

        updated_lines.append(updated_line)

    return "".join(updated_lines)


def _rename_signal_in_chart_view_state(old_name: str, new_name: str) -> list[Path]:
    if not old_name or not new_name or old_name == new_name:
        return []
    state = load_chart_view_state()
    changed = False
    for sheet in state.get("sheets", []):
        if not isinstance(sheet, dict):
            continue
        for panel in sheet.get("panels", []):
            if not isinstance(panel, dict):
                continue
            signals = panel.get("signals", [])
            if not isinstance(signals, list):
                continue
            renamed_signals = [new_name if str(signal).strip() == old_name else str(signal).strip() for signal in signals]
            if renamed_signals != signals:
                panel["signals"] = renamed_signals
                changed = True
    if not changed:
        return []
    save_chart_view_state(state)
    return [CHART_VIEW_STATE_PATH]


def rename_kpi_references(old_name: str, new_name: str) -> list[Path]:
    if not old_name or not new_name or old_name == new_name:
        return []

    updated_paths: list[Path] = []
    raw = _read_json(KPI_GROUPS_PATH)
    groups = raw.get("groups", []) if isinstance(raw, dict) else []
    changed_groups = False
    if isinstance(groups, list):
        for group in groups:
            if not isinstance(group, dict):
                continue
            kpis = group.get("kpis", [])
            if not isinstance(kpis, list):
                continue
            renamed_kpis = [new_name if str(kpi).strip() == old_name else str(kpi).strip() for kpi in kpis]
            if renamed_kpis != kpis:
                group["kpis"] = renamed_kpis
                changed_groups = True
    if changed_groups:
        KPI_GROUPS_PATH.write_text(json.dumps({"groups": groups}, ensure_ascii=False, indent=2), encoding="utf-8")
        updated_paths.append(KPI_GROUPS_PATH)

    updated_paths.extend(_rename_signal_in_chart_view_state(old_name, new_name))
    return updated_paths


def rename_derived_signal_references(old_name: str, new_name: str) -> list[Path]:
    if not old_name or not new_name or old_name == new_name:
        return []

    updated_paths: list[Path] = []
    for directory in [KPI_SPECS_DIR, DERIVED_SIGNALS_DIR]:
        for path in _iter_python_files(directory):
            if path.name == "00_example_and_guide.py":
                continue
            content = path.read_text(encoding="utf-8")
            renamed = _rename_identifier_references_in_python_config(content, old_name, new_name)
            if renamed != content:
                path.write_text(renamed, encoding="utf-8")
                updated_paths.append(path)

    updated_paths.extend(_rename_signal_in_chart_view_state(old_name, new_name))
    return updated_paths


def create_kpi_draft_file() -> Path:
    _ensure_supporting_files()
    target = _next_available_config_path(KPI_SPECS_DIR, "new_kpi", ".py")
    target.write_text(
        DEFAULT_KPI_TEMPLATE.format(kpi_name=target.stem, title="请填写中文KPI标题"),
        encoding="utf-8",
    )
    return target


def create_derived_signal_spec_file(display_name: str, signal_name: str | None = None) -> Path:
    _ensure_supporting_files()
    safe_stem = _slugify_file_stem(signal_name or display_name, "new_derived_signal")
    target = DERIVED_SIGNALS_DIR / f"{safe_stem}.py"
    if target.exists():
        raise FileExistsError(f"派生量文件已存在: {target.name}")
    final_signal_name = signal_name or safe_stem
    target.write_text(DEFAULT_DERIVED_SIGNAL_TEMPLATE.format(signal_name=final_signal_name, title=display_name), encoding="utf-8")
    return target


def create_derived_signal_draft_file(signal_name: str | None = None) -> Path:
    _ensure_supporting_files()
    target = (
        _next_available_config_path(DERIVED_SIGNALS_DIR, _slugify_file_stem(signal_name, "new_derived_signal"), ".py")
        if signal_name
        else _next_available_config_path(DERIVED_SIGNALS_DIR, "new_derived_signal", ".py")
    )
    target.write_text(
        DEFAULT_DERIVED_SIGNAL_TEMPLATE.format(signal_name=signal_name or target.stem, title="请填写中文派生量标题"),
        encoding="utf-8",
    )
    return target


def _normalize_chart_view_panels(raw_panels: Any) -> list[dict[str, Any]]:
    normalized_panels: list[dict[str, Any]] = []
    if not isinstance(raw_panels, list):
        return normalized_panels
    for item in raw_panels:
        if not isinstance(item, dict):
            continue
        signals = item.get("signals", [])
        if not isinstance(signals, list):
            continue
        normalized_panels.append({"signals": [str(signal).strip() for signal in signals if str(signal).strip()]})
    return normalized_panels


def _normalize_chart_view_state(raw: Any) -> dict[str, Any]:
    if not isinstance(raw, dict):
        return {"active_sheet": 0, "sheets": []}

    normalized_sheets: list[dict[str, Any]] = []
    raw_sheets = raw.get("sheets", [])
    if isinstance(raw_sheets, list):
        for index, item in enumerate(raw_sheets, start=1):
            if not isinstance(item, dict):
                continue
            name = str(item.get("name", "")).strip() or f"工作表 {index}"
            normalized_sheets.append({"name": name, "panels": _normalize_chart_view_panels(item.get("panels", []))})

    if not normalized_sheets:
        legacy_panels = _normalize_chart_view_panels(raw.get("panels", []))
        if legacy_panels:
            normalized_sheets.append({"name": "工作表 1", "panels": legacy_panels})

    active_sheet = raw.get("active_sheet", 0)
    if not isinstance(active_sheet, int):
        active_sheet = 0
    if normalized_sheets:
        active_sheet = max(0, min(active_sheet, len(normalized_sheets) - 1))
    else:
        active_sheet = 0

    return {"active_sheet": active_sheet, "sheets": normalized_sheets}


def load_chart_view_state() -> dict[str, Any]:
    _ensure_supporting_files()
    try:
        raw = json.loads(CHART_VIEW_STATE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {"active_sheet": 0, "sheets": []}
    return _normalize_chart_view_state(raw)


def save_chart_view_state(state: dict[str, Any]) -> Path:
    _ensure_supporting_files()
    payload = _normalize_chart_view_state(state)
    CHART_VIEW_STATE_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return CHART_VIEW_STATE_PATH


def load_formula_signal_definitions() -> list[dict[str, str]]:
    _ensure_supporting_files()
    raw = json.loads(FORMULA_SIGNALS_PATH.read_text(encoding="utf-8"))
    signals = raw.get("signals", []) if isinstance(raw, dict) else []
    definitions: list[dict[str, str]] = []
    for item in signals:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name", "")).strip()
        expression = str(item.get("expression", "")).strip()
        if name and expression:
            definitions.append({"name": name, "expression": expression})
    return sorted(definitions, key=lambda entry: entry["name"].lower())


def save_formula_signal_definition(name: str, expression: str) -> Path:
    _ensure_supporting_files()
    definitions = load_formula_signal_definitions()
    normalized_name = name.strip()
    normalized_expression = expression.strip()
    definitions = [item for item in definitions if item["name"] != normalized_name]
    definitions.append({"name": normalized_name, "expression": normalized_expression})
    FORMULA_SIGNALS_PATH.write_text(json.dumps({"signals": definitions}, ensure_ascii=False, indent=2), encoding="utf-8")
    return FORMULA_SIGNALS_PATH


def delete_formula_signal_definition(name: str) -> Path:
    _ensure_supporting_files()
    definitions = [item for item in load_formula_signal_definitions() if item["name"] != name]
    FORMULA_SIGNALS_PATH.write_text(json.dumps({"signals": definitions}, ensure_ascii=False, indent=2), encoding="utf-8")
    return FORMULA_SIGNALS_PATH


def create_report_template_file(display_name: str) -> Path:
    _ensure_supporting_files()
    safe_stem = _slugify_file_stem(display_name, "custom_report")
    target = REPORT_TEMPLATES_DIR / f"{safe_stem}.html"
    if target.exists():
        raise FileExistsError(f"模板文件已存在: {target.name}")
    target.write_text(DEFAULT_REPORT_TEMPLATE, encoding="utf-8")
    return target


def delete_config_file(path: str | Path) -> None:
    target = Path(path)
    if target.exists():
        target.unlink()


def _read_legacy_interface_mapping() -> dict[str, dict[str, Any]]:
    if not LEGACY_INTERFACE_MAPPING_PATH.exists():
        return {}

    raw = _read_json(LEGACY_INTERFACE_MAPPING_PATH)
    entries = raw.get("signals", []) if isinstance(raw, dict) else []
    mapping: dict[str, dict[str, Any]] = {}
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        signal_name = str(entry.get("standard_signal", "")).strip()
        if not signal_name:
            continue
        manual_column = str(entry.get("manual_column", "")).strip()
        mapping[signal_name] = {
            "actual_names": [manual_column] if manual_column else [],
            "required_by": [str(owner) for owner in entry.get("required_by", [])],
            "source_sheet": SYSTEM_MAPPING_SHEET,
        }
    return mapping


def _normalize_actual_names(values: list[Any]) -> list[str]:
    return [str(value).strip() for value in values if value is not None and str(value).strip()]


def load_interface_signal_tables() -> dict[str, list[dict[str, Any]]]:
    if not INTERFACE_MAPPING_PATH.exists():
        return {"system": [], "custom": []}
    workbook = load_workbook(INTERFACE_MAPPING_PATH)
    system_sheet = workbook[SYSTEM_MAPPING_SHEET] if SYSTEM_MAPPING_SHEET in workbook.sheetnames else workbook.active
    custom_sheet = workbook[CUSTOM_MAPPING_SHEET] if CUSTOM_MAPPING_SHEET in workbook.sheetnames else None
    reference_sheet = workbook[REFERENCE_SHEET] if REFERENCE_SHEET in workbook.sheetnames else None

    reference_lookup: dict[str, dict[str, Any]] = {}
    if reference_sheet is not None:
        for row in reference_sheet.iter_rows(min_row=2, values_only=True):
            signal_name = str(row[0] or "").strip()
            if not signal_name:
                continue
            reference_lookup[signal_name] = {
                "required_by": [item.strip() for item in str(row[1] or "").split("|") if item and item.strip()],
            }

    def read_sheet(sheet, sheet_name: str) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        if sheet is None:
            return rows
        header_row = [str(cell or "").strip().lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])]
        has_description_column = sheet_name == SYSTEM_MAPPING_SHEET and len(header_row) > 1 and header_row[1] == "description"
        has_from_column = sheet_name == SYSTEM_MAPPING_SHEET and ((len(header_row) > 2 and header_row[2] == "from") or (len(header_row) > 1 and header_row[1] == "from"))
        actual_name_start = 3 if has_description_column and has_from_column else (2 if has_from_column else 1)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            signal_name = str(row[0] or "").strip()
            if not signal_name:
                continue
            reference = reference_lookup.get(signal_name, {})
            rows.append(
                {
                    "standard_signal": signal_name,
                    "description": RAW_INPUT_DESCRIPTIONS.get(signal_name, str(row[1] or "").strip() if has_description_column else ""),
                    "actual_names": _normalize_actual_names(list(row[actual_name_start:])),
                    "required_by": reference.get("required_by", ["user_defined"] if sheet_name == CUSTOM_MAPPING_SHEET else []),
                    "source_sheet": sheet_name,
                }
            )
        return rows

    return {
        "system": read_sheet(system_sheet, SYSTEM_MAPPING_SHEET),
        "custom": read_sheet(custom_sheet, CUSTOM_MAPPING_SHEET),
    }


def _auto_fit_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        max_length = max((len(value) for value in values), default=0)
        column_letter = column_cells[0].column_letter
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def _protect_readonly_sheet(worksheet) -> None:
    worksheet.protection.sheet = True
    worksheet.protection.formatCells = True
    worksheet.protection.formatColumns = True
    worksheet.protection.formatRows = True
    worksheet.protection.insertColumns = False
    worksheet.protection.insertRows = False
    worksheet.protection.deleteColumns = False
    worksheet.protection.deleteRows = False


def _collect_runtime_raw_inputs() -> dict[str, list[str]]:
    required_by = collect_kpi_raw_input_requirements()
    for signal_name, owners in collect_derived_signal_raw_input_requirements().items():
        required_by.setdefault(signal_name, []).extend(owners)
    required_by.setdefault("time_s", [])
    return required_by


def save_interface_signal_tables(system_rows: list[dict[str, Any]], custom_rows: list[dict[str, Any]], actual_name_column_count: int = DEFAULT_INTERFACE_ACTUAL_NAME_COLUMN_COUNT) -> Path:
    required_by = _collect_runtime_raw_inputs()
    actual_name_column_count = max(DEFAULT_INTERFACE_ACTUAL_NAME_COLUMN_COUNT, int(actual_name_column_count or DEFAULT_INTERFACE_ACTUAL_NAME_COLUMN_COUNT))
    workbook = Workbook()
    system_sheet = workbook.active
    system_sheet.title = SYSTEM_MAPPING_SHEET
    custom_sheet = workbook.create_sheet(CUSTOM_MAPPING_SHEET)
    guide_sheet = workbook.create_sheet(GUIDE_SHEET)
    reference_sheet = workbook.create_sheet(REFERENCE_SHEET)
    metadata_sheet = workbook.create_sheet(INTERFACE_MAPPING_METADATA_SHEET)
    metadata_sheet.sheet_state = "hidden"
    metadata_sheet.append(["actual_signal_name_column_count", actual_name_column_count])

    guide_rows = [
        ["用途", "第一个 sheet 为系统自动维护的 KPI 与派生量 raw_inputs 信号，第二个 sheet 为用户自定义扩展信号。"],
        ["编辑规则", f"{SYSTEM_MAPPING_SHEET} 的 A-C 列只读，后续真实信号名列可编辑；{CUSTOM_MAPPING_SHEET} 的第一列和真实信号名列均可编辑。"],
        ["匹配逻辑", "分析时会按 actual_signal_name_1 到当前最后一列依次匹配；单元格里既可以填实际信号名，也可以填公式表达式，例如 time_ms*0.001。系统信号的 Description 列自动展示物理意义与单位，From 列用于说明该 raw_inputs 名称被哪些 KPI / 派生量引用。"],
        ["绘图信号", "曲线页下拉菜单会读取前两个 sheet 的第一列，因此自定义信号可用于绘图。"],
        ["注意事项", "自定义信号不会自动参与规则分析，除非你在规则/KPI Python 文件中显式引用它。"],
    ]
    for row in guide_rows:
        guide_sheet.append(row)
    _auto_fit_worksheet(guide_sheet)
    _protect_readonly_sheet(guide_sheet)

    def append_rows(sheet, rows: list[dict[str, Any]], headers: list[str], read_only_columns: set[int]) -> None:
        sheet.append(headers)
        for entry in rows:
            signal_name = str(entry.get("standard_signal", "")).strip()
            if not signal_name:
                continue
            actual_names = _normalize_actual_names(list(entry.get("actual_names", [])))[:actual_name_column_count]
            row = [signal_name]
            if headers[:3] == ["raw_input_name", "description", "from"]:
                row.append(RAW_INPUT_DESCRIPTIONS.get(signal_name, str(entry.get("description", "")).strip() or "请补充该输入量的物理意义与单位说明。"))
                row.append("\n".join(_format_required_by_owner(owner) for owner in entry.get("required_by", [])))
            row.extend(actual_names)
            while len(row) < len(headers):
                row.append("")
            sheet.append(row)
        for row_index in range(2, sheet.max_row + 1):
            for column_index in range(1, len(headers) + 1):
                sheet.cell(row=row_index, column=column_index).protection = Protection(locked=(column_index - 1) in read_only_columns)
        sheet.protection.sheet = True
        sheet.protection.formatColumns = True
        sheet.protection.formatRows = True
        sheet.protection.insertRows = not bool(read_only_columns)
        sheet.protection.deleteRows = not bool(read_only_columns)
        _auto_fit_worksheet(sheet)

    normalized_system_rows = []
    for signal_name in _ordered_raw_input_names(list(required_by)):
        existing = next((row for row in system_rows if str(row.get("standard_signal", "")).strip() == signal_name), None)
        normalized_system_rows.append({
            "standard_signal": signal_name,
            "actual_names": [] if existing is None else existing.get("actual_names", []),
            "description": RAW_INPUT_DESCRIPTIONS.get(signal_name, "请补充该输入量的物理意义与单位说明。"),
            "required_by": sorted(set(required_by.get(signal_name, []))),
        })
    append_rows(system_sheet, normalized_system_rows, build_system_interface_mapping_headers(actual_name_column_count), read_only_columns={0, 1, 2})

    filtered_custom_rows = []
    seen_custom: set[str] = set()
    for row in custom_rows:
        signal_name = str(row.get("standard_signal", "")).strip()
        if not signal_name or signal_name in seen_custom:
            continue
        filtered_custom_rows.append({"standard_signal": signal_name, "actual_names": row.get("actual_names", [])})
        seen_custom.add(signal_name)
    append_rows(custom_sheet, filtered_custom_rows, build_custom_interface_mapping_headers(actual_name_column_count), read_only_columns=set())

    reference_sheet.append(["raw_input_name", "from"])
    for signal_name in _ordered_raw_input_names(list(required_by)):
        reference_sheet.append([
            signal_name,
            " | ".join(sorted(set(required_by[signal_name]))),
        ])
    for row in filtered_custom_rows:
        signal_name = str(row["standard_signal"])
        reference_sheet.append([signal_name, "user_defined"])
    _auto_fit_worksheet(reference_sheet)
    _protect_readonly_sheet(reference_sheet)

    workbook.save(INTERFACE_MAPPING_PATH)
    return INTERFACE_MAPPING_PATH


def sync_interface_mapping_file() -> Path:
    existing_tables = None
    if INTERFACE_MAPPING_PATH.exists():
        workbook = load_workbook(INTERFACE_MAPPING_PATH)
        if SYSTEM_MAPPING_SHEET in workbook.sheetnames and CUSTOM_MAPPING_SHEET in workbook.sheetnames:
            existing_tables = load_interface_signal_tables()
    existing_entries = _read_interface_mapping_workbook() if existing_tables is not None else _read_legacy_interface_mapping()
    actual_name_column_count = get_interface_mapping_actual_name_column_count() if existing_tables is not None else DEFAULT_INTERFACE_ACTUAL_NAME_COLUMN_COUNT

    system_rows = []
    required_by = _collect_runtime_raw_inputs()
    for signal_name in _ordered_raw_input_names(list(required_by)):
        existing = existing_entries.get(signal_name, {})
        system_rows.append({"standard_signal": signal_name, "actual_names": existing.get("actual_names", [])})

    custom_rows = [] if existing_tables is None else existing_tables.get("custom", [])
    return save_interface_signal_tables(system_rows, custom_rows, actual_name_column_count=actual_name_column_count)


def _read_interface_mapping_workbook() -> dict[str, dict[str, Any]]:
    if not INTERFACE_MAPPING_PATH.exists():
        return {}
    workbook = load_workbook(INTERFACE_MAPPING_PATH)
    if SYSTEM_MAPPING_SHEET not in workbook.sheetnames:
        return _read_legacy_interface_mapping()
    tables = load_interface_signal_tables()
    mapping: dict[str, dict[str, Any]] = {}
    for row in [*tables.get("system", []), *tables.get("custom", [])]:
        signal_name = str(row.get("standard_signal", "")).strip()
        if not signal_name:
            continue
        mapping[signal_name] = {
            "actual_names": list(row.get("actual_names", [])),
            "required_by": list(row.get("required_by", [])),
            "source_sheet": row.get("source_sheet", SYSTEM_MAPPING_SHEET),
        }
    return mapping


def get_plot_signal_names() -> list[str]:
    tables = load_interface_signal_tables()
    names = {
        str(row.get("standard_signal", "")).strip()
        for row in [*tables.get("system", []), *tables.get("custom", [])]
        if str(row.get("standard_signal", "")).strip()
    }
    for item in load_formula_signal_definitions():
        names.add(item["name"])
    return sorted(names)


def load_interface_mapping() -> dict[str, dict[str, Any]]:
    sync_interface_mapping_file()
    mapping = _read_interface_mapping_workbook()
    for entry in mapping.values():
        actual_names = entry.get("actual_names", [])
        entry["manual_column"] = actual_names[0] if actual_names else ""
    return mapping
