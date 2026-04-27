import sys
import os
import pandas as pd
import openpyxl

sys.path.append(os.path.join(os.getcwd(), "src"))

from tcs_smart_analyzer.data.loaders import load_timeseries_file
from tcs_smart_analyzer.core.engine import AnalysisEngine

def task():
    file_path = r"tests/D档驱制动，10kph以内纯液压.xlsx"
    mapping_xlsx = r"src/tcs_smart_analyzer/config/interface_mapping.xlsx"

    # 1) Load file and check columns
    df = load_timeseries_file(file_path)
    cols = df.columns.tolist()
    results_1 = {c: (c in cols) for c in ["time", "time1", "time2", "time3"]}
    print(f"Columns check: {results_1}")

    # Load mapping workbook
    wb = openpyxl.load_workbook(mapping_xlsx)
    ws = wb.active 
    
    time_s_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "time_s":
            time_s_row = row
            break
    
    if not time_s_row:
        print("ERROR: \"time_s\" row not found in mapping file")
        return

    original_val = ws.cell(row=time_s_row, column=2).value
    engine = AnalysisEngine()

    for target in ["time1", "time2", "time3"]:
        ws.cell(row=time_s_row, column=2).value = target
        wb.save(mapping_xlsx)
        
        print(f"Testing with mapping time_s -> {target}:")
        try:
            engine.analyze_file(file_path)
            print("PASS")
        except Exception as e:
            print(f"ERROR: {e}")
            
    ws.cell(row=time_s_row, column=2).value = original_val
    wb.save(mapping_xlsx)
    print("Mapping restored.")

if __name__ == "__main__":
    task()
