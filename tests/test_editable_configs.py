from __future__ import annotations

import json
import unittest
from pathlib import Path

from openpyxl import load_workbook

from tcs_smart_analyzer.config.editable_configs import (
    align_python_config_file_name,
    create_derived_signal_draft_file,
    create_kpi_draft_file,
    delete_kpi_group,
    extract_derived_signal_name_from_text,
    extract_kpi_name_from_text,
    get_config_file_paths,
    load_derived_signal_definitions,
    load_formula_signal_definitions,
    load_chart_view_state,
    load_interface_signal_tables,
    load_kpi_definitions,
    load_kpi_groups,
    load_rule_definitions,
    list_derived_signal_spec_entries,
    list_kpi_spec_entries,
    list_rule_spec_entries,
    read_text_config_file,
    rename_derived_signal_references,
    save_chart_view_state,
    save_kpi_group,
    sync_interface_mapping_file,
    sync_legacy_deprecated_configs,
    validate_python_config_content,
    validate_runtime_definition_files,
)


class EditableConfigTests(unittest.TestCase):
    def test_kpi_definitions_are_available_and_rules_are_retired(self) -> None:
        rules = load_rule_definitions()
        kpis = load_kpi_definitions()
        kpi_names = {item["name"] for item in kpis}
        derived_names = {item["name"] for item in load_derived_signal_definitions()}
        baseline_kpis = {"max_jerk_mps3", "max_slip_speed", "mean_vehicle_speed_kph", "peak_slip_ratio"}

        self.assertEqual(len(rules), 0)
        self.assertTrue(baseline_kpis.issubset(kpi_names))
        self.assertIn("slip_ratio", derived_names)

    def test_rule_and_kpi_template_entries_are_first(self) -> None:
        rule_entries = list_rule_spec_entries()
        kpi_entries = list_kpi_spec_entries()
        derived_entries = list_derived_signal_spec_entries()

        self.assertTrue(rule_entries[0]["is_template"])
        self.assertTrue(kpi_entries[0]["is_template"])
        self.assertTrue(derived_entries[0]["is_template"])
        self.assertIn("示例", str(rule_entries[0]["display_name"]))
        self.assertIn("示例", str(kpi_entries[0]["display_name"]))
        self.assertIn("示例", str(derived_entries[0]["display_name"]))
        self.assertTrue(any(entry["display_name"] == "peak_slip_ratio 峰值打滑率" for entry in kpi_entries))
        self.assertTrue(any(entry["display_name"] == "slip_ratio 打滑率" for entry in derived_entries))

    def test_formula_signal_config_file_is_available(self) -> None:
        definitions = load_formula_signal_definitions()
        self.assertIsInstance(definitions, list)

    def test_interface_mapping_file_syncs_required_signals(self) -> None:
        path = sync_interface_mapping_file()
        workbook = load_workbook(path)
        mapping_sheet = workbook["系统信号"]
        custom_sheet = workbook["自定义信号"]
        reference_sheet = workbook["参考信息"]
        mapping_values = [mapping_sheet.cell(row=row_index, column=1).value for row_index in range(2, mapping_sheet.max_row + 1)]
        mapping_sources = [str(mapping_sheet.cell(row=row_index, column=2).value or "") for row_index in range(2, mapping_sheet.max_row + 1)]
        reference_values = [reference_sheet.cell(row=row_index, column=1).value for row_index in range(2, reference_sheet.max_row + 1)]

        self.assertIn("vehicle_speed", mapping_values)
        self.assertIn("longitudinal_accel_mps2", mapping_values)
        self.assertNotIn("brake_pressure_fl_bar", mapping_values)
        self.assertNotIn("brake_pressure_fr_bar", mapping_values)
        self.assertNotIn("torque_actual_nm", mapping_values)
        self.assertIn("vehicle_speed", reference_values)
        self.assertEqual(mapping_sheet.cell(row=1, column=2).value, "from")
        self.assertEqual(mapping_sheet.cell(row=1, column=1).value, "raw_input_name")
        self.assertEqual(reference_sheet.cell(row=1, column=1).value, "raw_input_name")
        self.assertEqual(reference_sheet.cell(row=1, column=2).value, "from")
        self.assertTrue(any("KPI:" in value or "派生量:" in value for value in mapping_sources))
        self.assertEqual(custom_sheet.max_column, 6)
        self.assertGreater(float(mapping_sheet.column_dimensions["A"].width), 10.0)

    def test_interface_mapping_file_removes_stale_system_signals(self) -> None:
        path = sync_interface_mapping_file()
        workbook = load_workbook(path)
        mapping_sheet = workbook["系统信号"]
        mapping_sheet.append(["torque_actual_nm", "legacy_signal", "", "", "", ""])
        workbook.save(path)

        synced_path = sync_interface_mapping_file()
        synced_workbook = load_workbook(synced_path)
        synced_values = [synced_workbook["系统信号"].cell(row=row_index, column=1).value for row_index in range(2, synced_workbook["系统信号"].max_row + 1)]

        self.assertNotIn("torque_actual_nm", synced_values)

    def test_interface_signal_tables_expose_system_and_custom_rows(self) -> None:
        sync_interface_mapping_file()
        tables = load_interface_signal_tables()

        self.assertIn("system", tables)
        self.assertIn("custom", tables)
        self.assertTrue(any(row["standard_signal"] == "vehicle_speed" for row in tables["system"]))
        self.assertTrue(any(row["required_by"] for row in tables["system"]))

    def test_derived_signal_definitions_include_algorithm_summary(self) -> None:
        definitions = load_derived_signal_definitions()
        slip_ratio = next(item for item in definitions if item["name"] == "slip_ratio")

        self.assertIn("algorithm_summary", slip_ratio)
        self.assertIn("vehicle_speed", str(slip_ratio["algorithm_summary"]))

    def test_kpi_guide_lists_current_derived_signals(self) -> None:
        config_paths = get_config_file_paths()
        guide_text = read_text_config_file(config_paths["kpi_specs_dir"] / "00_example_and_guide.py")
        derived_guide_text = read_text_config_file(config_paths["derived_signals_dir"] / "00_example_and_guide.py")

        self.assertIn("当前可用派生量清单", guide_text)
        self.assertIn("slip_ratio", guide_text)
        self.assertIn("当前可用派生量清单", derived_guide_text)
        self.assertIn("tcs_target_slip_ratio_global", derived_guide_text)
        self.assertIn("算法概述", guide_text)
        self.assertIn("算法概述", derived_guide_text)
        self.assertIn("唯一需要维护的输入声明", guide_text)
        self.assertIn("raw_inputs 就是唯一需要维护的输入声明", derived_guide_text)
        self.assertIn("派生量不需要 trend_source", derived_guide_text)
        self.assertIn("calculate_signal(dataframe) 的返回值本身就是曲线来源", derived_guide_text)
        self.assertIn("没有单独的 trend_source 字段", derived_guide_text)
        self.assertIn("DERIVED_SIGNAL_DEFINITION[\"name\"]", derived_guide_text)
        self.assertIn("calculate_kpi_series(dataframe)：必填字段", guide_text)
        self.assertIn("trend_source 不是可选项", guide_text)
        self.assertIn("必须使用英文", guide_text)
        self.assertIn("请使用中文", guide_text)
        self.assertIn("必须使用英文", derived_guide_text)
        self.assertIn("请使用中文", derived_guide_text)
        self.assertNotIn("DISPLAY_NAME = \"KPI 展示名称\"", guide_text)
        self.assertNotIn("DISPLAY_NAME = \"派生量展示名称\"", derived_guide_text)

    def test_kpi_runtime_contract_validation_requires_series_and_matching_trend_source(self) -> None:
        issues = validate_python_config_content(
            Path("demo_kpi.py"),
            '''KPI_DEFINITION = {
    "name": "demo_kpi",
    "trend_source": "other_series",
}

def calculate_kpi(dataframe):
    return 0.0
''',
        )

        self.assertTrue(any(issue.code == "kpi-missing-series" for issue in issues))
        self.assertTrue(any(issue.code == "kpi-trend-source-mismatch" for issue in issues))

    def test_python_config_validator_reports_syntax_and_undefined_name(self) -> None:
        issues = validate_python_config_content(Path("demo_kpi.py"), "def calculate_kpi(dataframe):\n    return missing_value\n")
        syntax_issues = validate_python_config_content(Path("demo_kpi.py"), "def broken(:\n    pass\n")

        self.assertTrue(any(issue.code == "undefined-name" and issue.message == "未定义名称: missing_value" for issue in issues))
        self.assertTrue(any(issue.code == "syntax-error" for issue in syntax_issues))

    def test_python_config_validator_does_not_misreport_loop_local_names(self) -> None:
        config_paths = get_config_file_paths()
        slip_ratio_path = config_paths["derived_signals_dir"] / "slip_ratio.py"
        issues = validate_python_config_content(slip_ratio_path, read_text_config_file(slip_ratio_path))

        self.assertEqual(issues, [])

    def test_chart_view_state_migrates_legacy_panels_to_default_sheet(self) -> None:
        config_paths = get_config_file_paths()
        chart_state_path = config_paths["chart_view_state"]
        original_text = chart_state_path.read_text(encoding="utf-8")
        chart_state_path.write_text(json.dumps({"panels": [{"signals": ["vehicle_speed"]}]}, ensure_ascii=False, indent=2), encoding="utf-8")
        try:
            state = load_chart_view_state()

            self.assertEqual(state["active_sheet"], 0)
            self.assertEqual(len(state["sheets"]), 1)
            self.assertEqual(state["sheets"][0]["name"], "工作表 1")
            self.assertEqual(state["sheets"][0]["panels"][0]["signals"], ["vehicle_speed"])
        finally:
            chart_state_path.write_text(original_text, encoding="utf-8")

    def test_save_chart_view_state_persists_sheet_names_and_active_index(self) -> None:
        config_paths = get_config_file_paths()
        chart_state_path = config_paths["chart_view_state"]
        original_text = chart_state_path.read_text(encoding="utf-8")
        try:
            save_chart_view_state(
                {
                    "active_sheet": 1,
                    "sheets": [
                        {"name": "总览", "panels": [{"signals": ["vehicle_speed"]}]},
                        {"name": "制动", "panels": [{"signals": ["wheel_speed_fl", "wheel_speed_fr"]}]},
                    ],
                }
            )
            reloaded = load_chart_view_state()

            self.assertEqual(reloaded["active_sheet"], 1)
            self.assertEqual([sheet["name"] for sheet in reloaded["sheets"]], ["总览", "制动"])
            self.assertEqual(reloaded["sheets"][1]["panels"][0]["signals"], ["wheel_speed_fl", "wheel_speed_fr"])
        finally:
            chart_state_path.write_text(original_text, encoding="utf-8")

    def test_validate_runtime_definition_files_reports_missing_derived_dependency(self) -> None:
        config_paths = get_config_file_paths()
        temp_kpi_path = config_paths["kpi_specs_dir"] / "temp_missing_dep_kpi.py"
        temp_kpi_path.write_text(
            '''from __future__ import annotations

IS_TEMPLATE = False

KPI_DEFINITION = {
    "name": "temp_missing_dep_kpi",
    "title": "缺失依赖测试KPI",
    "raw_inputs": ["vehicle_speed"],
    "derived_inputs": ["missing_derived_signal_for_test"],
    "trend_source": "temp_missing_dep_kpi",
    "unit": "",
    "description": "测试缺失派生量依赖提示。",
    "threshold": 0.0,
    "source": "test",
    "pass_condition": "value >= threshold",
    "rule_description": "test",
    "pass_message": "pass",
    "fail_message": "fail",
}

def calculate_kpi(dataframe):
    return 0.0

def calculate_kpi_series(dataframe):
    return dataframe["vehicle_speed"]
''',
            encoding="utf-8",
        )
        try:
            issues = validate_runtime_definition_files()

            self.assertTrue(any("missing_derived_signal_for_test" in issue.message for issue in issues))
            self.assertTrue(any(Path(issue.path) == temp_kpi_path for issue in issues))
        finally:
            if temp_kpi_path.exists():
                temp_kpi_path.unlink()

    def test_align_python_config_file_name_renames_to_definition_name(self) -> None:
        config_paths = get_config_file_paths()
        temp_path = config_paths["derived_signals_dir"] / "temp_old_name.py"
        temp_path.write_text("IS_TEMPLATE = False\n", encoding="utf-8")
        try:
            renamed_path = align_python_config_file_name(temp_path, "renamed_signal")

            self.assertEqual(renamed_path.name, "renamed_signal.py")
            self.assertTrue(renamed_path.exists())
            self.assertFalse(temp_path.exists())
        finally:
            if temp_path.exists():
                temp_path.unlink()
            renamed = config_paths["derived_signals_dir"] / "renamed_signal.py"
            if renamed.exists():
                renamed.unlink()

    def test_new_kpi_draft_is_created_as_real_definition(self) -> None:
        path = create_kpi_draft_file()
        try:
            content = read_text_config_file(path)

            self.assertIn("IS_TEMPLATE = False", content)
            self.assertIn(f'"name": "{path.stem}"', content)
            self.assertIn(f'"trend_source": "{path.stem}"', content)
            self.assertEqual(extract_kpi_name_from_text(content), path.stem)
        finally:
            if path.exists():
                path.unlink()

    def test_loading_kpi_entries_auto_renames_misaligned_file_name(self) -> None:
        config_paths = get_config_file_paths()
        temp_path = config_paths["kpi_specs_dir"] / "temp_old_name.py"
        temp_path.write_text(
            '''from __future__ import annotations

IS_TEMPLATE = False

KPI_DEFINITION = {
    "name": "temp_runtime_kpi_name",
    "title": "临时运行时KPI",
    "raw_inputs": ["time_s", "vehicle_speed"],
    "derived_inputs": [],
    "trend_source": "temp_runtime_kpi_name",
    "unit": "",
    "description": "",
    "threshold": 0.0,
    "source": "manual",
    "pass_condition": "value <= threshold",
    "rule_description": "",
    "pass_message": "",
    "fail_message": "",
}

def calculate_kpi(dataframe):
    return 0.0

def calculate_kpi_series(dataframe):
    return dataframe["vehicle_speed"]
''',
            encoding="utf-8",
        )
        renamed_path = config_paths["kpi_specs_dir"] / "temp_runtime_kpi_name.py"
        try:
            entries = list_kpi_spec_entries()

            self.assertTrue(renamed_path.exists())
            self.assertFalse(temp_path.exists())
            self.assertTrue(any(str(entry["path"]).endswith("temp_runtime_kpi_name.py") for entry in entries))
        finally:
            if temp_path.exists():
                temp_path.unlink()
            if renamed_path.exists():
                renamed_path.unlink()

    def test_non_guide_kpi_file_is_not_treated_as_template_even_if_flagged(self) -> None:
        config_paths = get_config_file_paths()
        temp_path = config_paths["kpi_specs_dir"] / "temp_flagged_kpi.py"
        temp_path.write_text(
            '''from __future__ import annotations

IS_TEMPLATE = True

KPI_DEFINITION = {
    "name": "temp_flagged_kpi",
    "title": "临时测试KPI",
    "raw_inputs": ["vehicle_speed"],
    "derived_inputs": [],
    "trend_source": "temp_flagged_kpi",
    "unit": "unitless",
    "description": "test",
    "threshold": 0.0,
    "source": "manual",
    "pass_condition": "value <= threshold",
    "rule_description": "test",
    "pass_message": "pass",
    "fail_message": "fail",
}

def calculate_kpi(dataframe):
    return 0.0

def calculate_kpi_series(dataframe):
    return [0.0] * len(dataframe)
''',
            encoding="utf-8",
        )
        try:
            kpi_entries = list_kpi_spec_entries()
            self.assertTrue(any(str(entry["display_name"]) == "temp_flagged_kpi 临时测试KPI" for entry in kpi_entries))
            self.assertIn("temp_flagged_kpi", {item["name"] for item in load_kpi_definitions()})
        finally:
            if temp_path.exists():
                temp_path.unlink()

    def test_new_derived_signal_draft_is_created_as_real_definition(self) -> None:
        path = create_derived_signal_draft_file()
        try:
            content = read_text_config_file(path)

            self.assertIn("IS_TEMPLATE = False", content)
            self.assertIn(f'"name": "{path.stem}"', content)
            self.assertEqual(extract_derived_signal_name_from_text(content), path.stem)
        finally:
            if path.exists():
                path.unlink()

    def test_new_derived_signal_draft_uses_requested_signal_name_for_file_name(self) -> None:
        path = create_derived_signal_draft_file("custom_signal_name")
        try:
            content = read_text_config_file(path)

            self.assertEqual(path.name, "custom_signal_name.py")
            self.assertIn('"name": "custom_signal_name"', content)
        finally:
            if path.exists():
                path.unlink()

    def test_loading_derived_entries_auto_renames_misaligned_file_name(self) -> None:
        config_paths = get_config_file_paths()
        temp_path = config_paths["derived_signals_dir"] / "temp_old_name.py"
        temp_path.write_text(
            '''from __future__ import annotations

IS_TEMPLATE = False

import pandas as pd

DERIVED_SIGNAL_DEFINITION = {
    "name": "temp_runtime_signal_name",
    "title": "临时运行时派生量",
    "raw_inputs": ["vehicle_speed"],
    "derived_inputs": [],
    "description": "",
    "algorithm_summary": "",
}

def calculate_signal(dataframe):
    return pd.to_numeric(dataframe["vehicle_speed"], errors="coerce")
''',
            encoding="utf-8",
        )
        renamed_path = config_paths["derived_signals_dir"] / "temp_runtime_signal_name.py"
        try:
            entries = list_derived_signal_spec_entries()

            self.assertTrue(renamed_path.exists())
            self.assertFalse(temp_path.exists())
            self.assertTrue(any(str(entry["path"]).endswith("temp_runtime_signal_name.py") for entry in entries))
        finally:
            if temp_path.exists():
                temp_path.unlink()
            if renamed_path.exists():
                renamed_path.unlink()

    def test_renaming_derived_signal_updates_dependent_configs(self) -> None:
        config_paths = get_config_file_paths()
        peak_path = config_paths["kpi_specs_dir"] / "peak_slip_ratio.py"
        derived_path = config_paths["derived_signals_dir"] / "tcs_target_slip_ratio_global.py"
        chart_state_path = config_paths["chart_view_state"]
        peak_backup = peak_path.read_text(encoding="utf-8")
        derived_backup = derived_path.read_text(encoding="utf-8")
        chart_backup = chart_state_path.read_text(encoding="utf-8") if chart_state_path.exists() else None

        try:
            save_chart_view_state({"active_sheet": 0, "sheets": [{"name": "工作表 1", "panels": [{"signals": ["slip_ratio", "vehicle_speed"]}]}]})
            rename_derived_signal_references("slip_ratio", "slip_ratio_renamed")

            peak_text = peak_path.read_text(encoding="utf-8")
            derived_text = derived_path.read_text(encoding="utf-8")
            self.assertIn('"derived_inputs": ["slip_ratio_renamed"]', peak_text)
            self.assertIn('"trend_source": "peak_slip_ratio"', peak_text)
            self.assertIn('"derived_inputs": ["slip_ratio_renamed"]', derived_text)
            self.assertEqual(load_chart_view_state()["sheets"][0]["panels"][0]["signals"], ["slip_ratio_renamed", "vehicle_speed"])
        finally:
            peak_path.write_text(peak_backup, encoding="utf-8")
            derived_path.write_text(derived_backup, encoding="utf-8")
            if chart_backup is None:
                if chart_state_path.exists():
                    chart_state_path.unlink()
            else:
                chart_state_path.write_text(chart_backup, encoding="utf-8")

    def test_chart_view_state_can_be_saved_and_restored(self) -> None:
        config_paths = get_config_file_paths()
        state_path = config_paths["chart_view_state"]
        backup = state_path.read_text(encoding="utf-8") if state_path.exists() else None
        try:
            save_chart_view_state(
                {
                    "active_sheet": 0,
                    "sheets": [{"name": "总览", "panels": [{"signals": ["vehicle_speed", "slip_ratio"]}, {"signals": []}]}],
                }
            )
            state = load_chart_view_state()

            self.assertEqual(state["sheets"][0]["name"], "总览")
            self.assertEqual(state["sheets"][0]["panels"][0]["signals"], ["vehicle_speed", "slip_ratio"])
            self.assertEqual(state["sheets"][0]["panels"][1]["signals"], [])
        finally:
            if backup is not None:
                state_path.write_text(backup, encoding="utf-8")

    def test_legacy_json_files_are_marked_deprecated(self) -> None:
        sync_legacy_deprecated_configs()
        project_root = Path(__file__).resolve().parents[1]
        rule_json = json.loads((project_root / "src" / "tcs_smart_analyzer" / "config" / "rule_definitions.json").read_text(encoding="utf-8"))
        kpi_json = json.loads((project_root / "src" / "tcs_smart_analyzer" / "config" / "kpi_definitions.json").read_text(encoding="utf-8"))

        self.assertTrue(rule_json["deprecated"])
        self.assertTrue(kpi_json["deprecated"])

    def test_kpi_groups_can_be_saved_and_restored(self) -> None:
        config_paths = get_config_file_paths()
        group_path = config_paths["kpi_groups"]
        backup = group_path.read_text(encoding="utf-8") if group_path.exists() else None
        try:
            save_kpi_group("测试分组", ["peak_slip_ratio", "max_jerk_mps3"], key="test_group")
            groups = load_kpi_groups()

            self.assertEqual(groups[0]["key"], "__all_kpis__")
            self.assertTrue(any(group["key"] == "test_group" for group in groups))

            filtered = load_kpi_definitions("test_group")
            self.assertEqual({item["name"] for item in filtered}, {"peak_slip_ratio", "max_jerk_mps3"})

            delete_kpi_group("test_group")
            self.assertFalse(any(group["key"] == "test_group" for group in load_kpi_groups()))
        finally:
            if backup is not None:
                group_path.write_text(backup, encoding="utf-8")


if __name__ == "__main__":
    unittest.main()